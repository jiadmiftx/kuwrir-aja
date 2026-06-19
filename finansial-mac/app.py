from flask import (Flask, render_template, request, redirect, url_for,
                   flash, jsonify, Response, session, send_file)
import sqlite3
from datetime import datetime, date, timedelta
import json, os, io, calendar
from hashlib import sha256
from functools import wraps

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

# Resolve base paths — works in dev, --onedir, and --onefile PyInstaller
import sys as _sys
if getattr(_sys, 'frozen', False):
    _BUNDLE = _sys._MEIPASS                        # read-only temp extraction dir
    _DATA   = os.path.dirname(_sys.executable)     # next to EXE — writable, persistent
else:
    _BUNDLE = os.path.dirname(os.path.abspath(__file__))
    _DATA   = _BUNDLE

app = Flask(__name__,
    template_folder=os.path.join(_BUNDLE, 'templates'),
    static_folder=os.path.join(_BUNDLE, 'static'))
_key_file = os.path.join(_DATA, '.secret_key')
if not os.path.exists(_key_file):
    with open(_key_file, 'wb') as _f:
        _f.write(os.urandom(32))
with open(_key_file, 'rb') as _f:
    app.secret_key = _f.read()
DB = os.path.join(_DATA, 'finansial.db')
UPLOAD_FOLDER = os.path.join(_DATA, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_IMG = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'svg'}

# Serve user-uploaded files from persistent DATA dir (supports --onefile bundle)
from flask import send_from_directory as _send_from_dir
@app.route('/static/uploads/<path:filename>')
def _serve_upload(filename):
    return _send_from_dir(UPLOAD_FOLDER, filename)

BULAN = ['Jan','Feb','Mar','Apr','Mei','Jun','Jul','Agt','Sep','Okt','Nov','Des']

# ---------- DB ----------
def db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    conn = db()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY, value TEXT
    );
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        nama TEXT,
        role TEXT NOT NULL DEFAULT 'VIEWER',
        aktif INTEGER DEFAULT 1,
        dibuat DATETIME DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS akun (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        kode TEXT UNIQUE NOT NULL,
        nama TEXT NOT NULL,
        tipe TEXT NOT NULL,
        subtipe TEXT,
        saldo_normal TEXT NOT NULL DEFAULT 'DEBIT'
    );
    CREATE TABLE IF NOT EXISTS jurnal (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tanggal DATE NOT NULL,
        keterangan TEXT NOT NULL,
        referensi TEXT,
        kategori TEXT DEFAULT 'OPERASIONAL',
        tipe_tx TEXT DEFAULT 'JURNAL',
        dibuat DATETIME DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS detail_jurnal (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        jurnal_id INTEGER NOT NULL,
        akun_id INTEGER NOT NULL,
        debit REAL DEFAULT 0,
        kredit REAL DEFAULT 0,
        FOREIGN KEY (jurnal_id) REFERENCES jurnal(id) ON DELETE CASCADE,
        FOREIGN KEY (akun_id) REFERENCES akun(id)
    );
    CREATE TABLE IF NOT EXISTS produk (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        kode TEXT UNIQUE NOT NULL,
        nama TEXT NOT NULL,
        varian TEXT,
        satuan TEXT DEFAULT 'pcs',
        harga_beli REAL DEFAULT 0,
        harga_jual REAL DEFAULT 0,
        stok REAL DEFAULT 0,
        min_stok REAL DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS pergerakan_stok (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        produk_id INTEGER NOT NULL,
        tanggal DATE NOT NULL,
        jenis TEXT NOT NULL,
        qty REAL NOT NULL,
        harga REAL DEFAULT 0,
        keterangan TEXT,
        dibuat DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (produk_id) REFERENCES produk(id)
    );
    CREATE TABLE IF NOT EXISTS piutang (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tanggal DATE NOT NULL,
        jatuh_tempo DATE,
        pelanggan TEXT NOT NULL,
        keterangan TEXT,
        jumlah REAL NOT NULL,
        terbayar REAL DEFAULT 0,
        status TEXT DEFAULT 'BELUM LUNAS',
        dibuat DATETIME DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS hutang (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tanggal DATE NOT NULL,
        jatuh_tempo DATE,
        pemasok TEXT NOT NULL,
        keterangan TEXT,
        jumlah REAL NOT NULL,
        terbayar REAL DEFAULT 0,
        status TEXT DEFAULT 'BELUM LUNAS',
        dibuat DATETIME DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS bayar_piutang (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        piutang_id INTEGER NOT NULL,
        tanggal DATE NOT NULL,
        jumlah REAL NOT NULL,
        catatan TEXT,
        FOREIGN KEY (piutang_id) REFERENCES piutang(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS bayar_hutang (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hutang_id INTEGER NOT NULL,
        tanggal DATE NOT NULL,
        jumlah REAL NOT NULL,
        catatan TEXT,
        FOREIGN KEY (hutang_id) REFERENCES hutang(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS aset_tetap (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nama TEXT NOT NULL,
        kategori TEXT,
        harga_beli REAL NOT NULL,
        tanggal_beli DATE NOT NULL,
        masa_pakai INTEGER NOT NULL,
        penyusutan_bulan REAL,
        aktif INTEGER DEFAULT 1,
        dibuat DATETIME DEFAULT CURRENT_TIMESTAMP
    );
    """)
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS invoice (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nomor TEXT UNIQUE NOT NULL,
        tanggal DATE NOT NULL,
        jatuh_tempo DATE,
        top_hari INTEGER,
        top_note TEXT DEFAULT '',
        pelanggan TEXT NOT NULL,
        alamat_pelanggan TEXT DEFAULT '',
        telepon_pelanggan TEXT DEFAULT '',
        diskon REAL DEFAULT 0,
        ongkir REAL DEFAULT 0,
        biaya_lain REAL DEFAULT 0,
        catatan TEXT DEFAULT '',
        status TEXT DEFAULT 'DRAFT',
        dibuat DATETIME DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS invoice_item (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        invoice_id INTEGER NOT NULL,
        deskripsi TEXT NOT NULL,
        qty REAL DEFAULT 1,
        satuan TEXT DEFAULT 'pcs',
        harga_satuan REAL DEFAULT 0,
        diskon_item REAL DEFAULT 0,
        subtotal REAL DEFAULT 0,
        FOREIGN KEY (invoice_id) REFERENCES invoice(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS hpp_produk (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nama TEXT NOT NULL,
        harga_jual REAL DEFAULT 0,
        bahan TEXT DEFAULT '[]',
        updated_at TEXT DEFAULT (date('now'))
    );
    CREATE TABLE IF NOT EXISTS log_aktivitas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        waktu DATETIME DEFAULT CURRENT_TIMESTAMP,
        user_id INTEGER,
        username TEXT,
        aksi TEXT NOT NULL,
        detail TEXT
    );
    """)
    # migrations
    for col_sql in [
        "ALTER TABLE jurnal ADD COLUMN tipe_tx TEXT DEFAULT 'JURNAL'",
        "ALTER TABLE produk ADD COLUMN varian TEXT",
        "ALTER TABLE akun ADD COLUMN is_rekening INTEGER DEFAULT 0",
        "ALTER TABLE akun ADD COLUMN no_rekening TEXT",
        "ALTER TABLE aset_tetap ADD COLUMN bulan_penyusutan_dicatat INTEGER DEFAULT 0",
        "ALTER TABLE aset_tetap ADD COLUMN akumulasi_penyusutan REAL DEFAULT 0",
        "ALTER TABLE log_aktivitas ADD COLUMN kategori TEXT DEFAULT 'LAINNYA'",
        "ALTER TABLE produk ADD COLUMN kategori TEXT DEFAULT ''",
        "ALTER TABLE invoice ADD COLUMN top_hari INTEGER",
        "ALTER TABLE invoice ADD COLUMN top_note TEXT DEFAULT ''",
    ]:
        try: conn.execute(col_sql)
        except: pass
    conn.execute("UPDATE akun SET is_rekening=1 WHERE kode IN ('1100','1110')")

    if conn.execute('SELECT COUNT(*) FROM akun').fetchone()[0] == 0:
        akun_default = [
            ('1100','Kas','ASET','Aset Lancar','DEBIT'),
            ('1110','Bank','ASET','Aset Lancar','DEBIT'),
            ('1120','Piutang Usaha','ASET','Aset Lancar','DEBIT'),
            ('1130','Persediaan Barang','ASET','Aset Lancar','DEBIT'),
            ('1140','Biaya Dibayar Dimuka','ASET','Aset Lancar','DEBIT'),
            ('1200','Peralatan','ASET','Aset Tetap','DEBIT'),
            ('1210','Kendaraan','ASET','Aset Tetap','DEBIT'),
            ('1220','Gedung & Bangunan','ASET','Aset Tetap','DEBIT'),
            ('1290','Akumulasi Penyusutan','ASET','Aset Tetap','KREDIT'),
            ('2100','Hutang Usaha','LIABILITAS','Liabilitas Lancar','KREDIT'),
            ('2110','Hutang Pajak','LIABILITAS','Liabilitas Lancar','KREDIT'),
            ('2120','Hutang Gaji','LIABILITAS','Liabilitas Lancar','KREDIT'),
            ('2130','Pendapatan Diterima Dimuka','LIABILITAS','Liabilitas Lancar','KREDIT'),
            ('2200','Hutang Bank','LIABILITAS','Liabilitas Jangka Panjang','KREDIT'),
            ('3100','Modal Pemilik','EKUITAS','Ekuitas','KREDIT'),
            ('3200','Laba Ditahan','EKUITAS','Ekuitas','KREDIT'),
            ('3300','Prive / Penarikan Owner','EKUITAS','Ekuitas','DEBIT'),
            ('4100','Pendapatan Penjualan','PENDAPATAN','Pendapatan Usaha','KREDIT'),
            ('4150','Retur Penjualan','PENDAPATAN','Pendapatan Usaha','DEBIT'),
            ('4200','Pendapatan Jasa','PENDAPATAN','Pendapatan Usaha','KREDIT'),
            ('4300','Pendapatan Lain-lain','PENDAPATAN','Pendapatan Lain','KREDIT'),
            ('5100','Harga Pokok Penjualan','BEBAN','HPP','DEBIT'),
            ('5150','Retur Pembelian','BEBAN','HPP','KREDIT'),
            ('6100','Beban Gaji','BEBAN','Beban Operasional','DEBIT'),
            ('6110','Beban Sewa','BEBAN','Beban Operasional','DEBIT'),
            ('6120','Beban Listrik & Air','BEBAN','Beban Operasional','DEBIT'),
            ('6130','Beban Penyusutan','BEBAN','Beban Penyusutan','DEBIT'),
            ('6140','Beban Pemasaran','BEBAN','Beban Operasional','DEBIT'),
            ('6150','Beban Administrasi','BEBAN','Beban Operasional','DEBIT'),
            ('6160','Beban Bunga','BEBAN','Beban Bunga','DEBIT'),
            ('6170','Beban Pajak','BEBAN','Beban Pajak','DEBIT'),
            ('6180','Beban Lainnya','BEBAN','Beban Operasional','DEBIT'),
        ]
        conn.executemany('INSERT INTO akun(kode,nama,tipe,subtipe,saldo_normal) VALUES(?,?,?,?,?)', akun_default)

    if conn.execute('SELECT COUNT(*) FROM settings').fetchone()[0] == 0:
        conn.executemany('INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)', [
            ('account_type','single'), ('modal_awal','0'), ('nama_usaha','Usaha Saya'),
        ])

    if conn.execute('SELECT COUNT(*) FROM users').fetchone()[0] == 0:
        ph = sha256('admin123'.encode()).hexdigest()
        conn.execute("INSERT INTO users(username,password_hash,nama,role) VALUES(?,?,?,?)",
                     ('admin', ph, 'Administrator', 'ADMIN'))
    ph_demo = sha256('demo123'.encode()).hexdigest()
    conn.execute(
        "INSERT OR IGNORE INTO users(username,password_hash,nama,role) VALUES(?,?,?,?)",
        ('demo', ph_demo, 'Demo User', 'DEMO')
    )

    # Migrasi: selaraskan subtipe akun beban dengan pengelompokan calc_profitability
    conn.execute("UPDATE akun SET subtipe='Beban Penyusutan'  WHERE kode='6130' AND subtipe='Beban Operasional'")
    conn.execute("UPDATE akun SET subtipe='Beban Bunga'       WHERE kode='6160' AND subtipe IN ('Beban Lain','Beban Operasional')")
    conn.execute("UPDATE akun SET subtipe='Beban Pajak'       WHERE kode='6170' AND subtipe IN ('Beban Operasional','Beban Lain')")
    conn.execute("UPDATE akun SET subtipe='Beban Operasional' WHERE kode='6180' AND subtipe='Beban Lain'")
    # Migrasi: tambah akun yang mungkin belum ada di database lama
    missing = [
        ('6170', 'Beban Pajak',       'BEBAN',      'Beban Pajak',        'DEBIT'),
        ('6180', 'Beban Lainnya',     'BEBAN',      'Beban Operasional',  'DEBIT'),
        ('4150', 'Retur Penjualan',   'PENDAPATAN', 'Pendapatan Usaha',   'DEBIT'),
        ('5150', 'Retur Pembelian',   'BEBAN',      'HPP',                'KREDIT'),
    ]
    for kode, nama, tipe, subtipe, saldo_normal in missing:
        conn.execute(
            "INSERT OR IGNORE INTO akun(kode,nama,tipe,subtipe,saldo_normal) VALUES(?,?,?,?,?)",
            (kode, nama, tipe, subtipe, saldo_normal)
        )

    conn.commit(); conn.close()


def parse_rp(v):
    """Parse Indonesian-formatted number strings. '2.250.000' → 2250000.0"""
    if v is None or v == '': return 0.0
    return float(str(v).replace('.', '').replace(',', '.')) or 0.0

# ---------- FILTERS ----------
@app.template_filter('rp')
def rp_filter(v):
    if v is None: v = 0
    v = float(v)
    prefix = '-Rp' if v < 0 else 'Rp'
    return "{} {:,.0f}".format(prefix, abs(v)).replace(',', '.')

@app.template_filter('tgl')
def tgl_filter(v):
    if not v: return '-'
    try:
        d = datetime.strptime(str(v)[:10], '%Y-%m-%d')
        return f"{d.day} {BULAN[d.month-1]} {d.year}"
    except: return str(v)

@app.context_processor
def inject_globals():
    conn = db()
    nama_usaha = get_setting(conn, 'nama_usaha', 'FinansialApp')
    conn.close()
    return {'now': datetime.now(), 'today': date.today(),
            'session': session, 'nama_usaha': nama_usaha}


# ---------- AUTH ----------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login'))
        if session.get('role') != 'ADMIN':
            flash('Akses ditolak. Hanya Admin.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def investor_required(f):
    """ADMIN + FINANCE + INVESTOR — read-only financial views."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login'))
        if session.get('role') not in ('ADMIN', 'FINANCE', 'INVESTOR', 'DEMO'):
            flash('Akses ditolak. Halaman ini hanya untuk Admin, Finance, dan Investor.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def finance_required(f):
    """ADMIN + FINANCE — full financial write access."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login'))
        if session.get('role') not in ('ADMIN', 'FINANCE', 'DEMO'):
            flash('Akses ditolak. Halaman ini hanya untuk Admin dan Finance.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def operator_required(f):
    """ADMIN + FINANCE + OPERATOR — input pages."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login'))
        if session.get('role') not in ('ADMIN', 'FINANCE', 'OPERATOR', 'DEMO'):
            flash('Akses ditolak. Halaman ini hanya untuk Admin, Finance, dan Operator.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

@app.before_request
def demo_readonly():
    if session.get('role') != 'DEMO':
        return
    if request.method in ('POST', 'PUT', 'DELETE', 'PATCH'):
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from flask import jsonify
            return jsonify({'error': 'User Demo tidak bisa menggunakan fitur ini'}), 403
        flash('User Demo tidak bisa menggunakan fitur ini', 'warning')
        return redirect(request.referrer or url_for('dashboard'))


def _operator_date_ok(tanggal_str):
    """Return True if date is in current month, or user is not OPERATOR."""
    if session.get('role') != 'OPERATOR':
        return True
    try:
        tgl = datetime.strptime(tanggal_str, '%Y-%m-%d').date()
        today = date.today()
        return tgl.year == today.year and tgl.month == today.month
    except ValueError:
        return False

@app.route('/login', methods=['GET','POST'])
def login():
    if session.get('user_id'):
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        ph = sha256(password.encode()).hexdigest()
        conn = db()
        user = conn.execute(
            "SELECT * FROM users WHERE username=? AND password_hash=? AND aktif=1",
            (username, ph)
        ).fetchone()
        conn.close()
        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['nama'] = user['nama'] or user['username']
            session['role'] = user['role']
            conn2 = db()
            add_log(conn2, 'Login', f"Role: {user['role']}", 'ADMIN')
            conn2.commit(); conn2.close()
            return redirect(url_for('dashboard'))
        flash('Username atau password salah.', 'danger')
    return render_template('login.html')

@app.route('/ganti-password', methods=['POST'])
@admin_required
def ganti_password():
    uid = session['user_id']
    pw_lama    = request.form.get('pw_lama', '')
    pw_baru    = request.form.get('pw_baru', '')
    pw_konfirm = request.form.get('pw_konfirm', '')

    if len(pw_baru) < 6:
        flash('Password baru minimal 6 karakter.', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    if pw_baru != pw_konfirm:
        flash('Konfirmasi password baru tidak cocok.', 'danger')
        return redirect(request.referrer or url_for('dashboard'))

    conn = db()
    user = conn.execute("SELECT password_hash FROM users WHERE id=?", (uid,)).fetchone()
    if not user or sha256(pw_lama.encode()).hexdigest() != user['password_hash']:
        conn.close()
        flash('Password lama tidak benar.', 'danger')
        return redirect(request.referrer or url_for('dashboard'))

    conn.execute("UPDATE users SET password_hash=? WHERE id=?",
                 (sha256(pw_baru.encode()).hexdigest(), uid))
    add_log(conn, 'Ubah password sendiri', f"User: {session.get('username')}", 'PASSWORD')
    conn.commit()
    conn.close()
    flash('Password berhasil diubah!', 'success')
    return redirect(request.referrer or url_for('dashboard'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ---------- HELPERS ----------
def _qv(conn, where_sql, params):
    r = conn.execute(f"""
        SELECT COALESCE(SUM(
            CASE WHEN a.saldo_normal='DEBIT' THEN d.debit-d.kredit
                 ELSE d.kredit-d.debit END
        ), 0) as v
        FROM akun a
        JOIN detail_jurnal d ON d.akun_id=a.id
        JOIN jurnal j ON j.id=d.jurnal_id
        WHERE {where_sql}
    """, params).fetchone()
    return float(r['v']) if r else 0.0

def calc_profitability(conn, sd, ed):
    def q(where, p): return _qv(conn, where, p)
    rev      = q("a.tipe='PENDAPATAN' AND j.tanggal>=? AND j.tanggal<=?", [sd,ed])
    hpp      = q("a.kode IN ('5100','5150') AND j.tanggal>=? AND j.tanggal<=?", [sd,ed])
    # Biaya operasional: semua beban kecuali HPP/retur-beli, penyusutan, bunga, dan pajak
    op_exp   = q("a.tipe='BEBAN' AND a.kode NOT IN ('5100','5150','6130','6160','6170') AND j.tanggal>=? AND j.tanggal<=?", [sd,ed])
    depr     = q("a.kode='6130' AND j.tanggal>=? AND j.tanggal<=?", [sd,ed])
    interest = q("a.kode='6160' AND j.tanggal>=? AND j.tanggal<=?", [sd,ed])
    tax      = q("a.kode='6170' AND j.tanggal>=? AND j.tanggal<=?", [sd,ed])
    prive    = q("a.kode='3300' AND j.tanggal>=? AND j.tanggal<=?", [sd,ed])

    laba_kotor  = rev - hpp
    laba_op     = laba_kotor - op_exp        # Margin Operasional (sebelum penyusutan)
    ebit        = laba_op - depr             # EBIT (setelah penyusutan, sebelum bunga & pajak)
    laba_bersih = ebit - interest - tax
    laba_tahan  = laba_bersih - prive

    def pct(v): return round(v/rev*100,1) if rev else 0
    return dict(
        rev=rev, hpp=hpp, op_exp=op_exp, depr=depr, interest=interest,
        tax=tax, prive=prive,
        laba_kotor=laba_kotor, laba_op=laba_op, ebit=ebit,
        laba_bersih=laba_bersih, laba_tahan=laba_tahan,
        other_e=interest,
        pct_hpp=pct(hpp), pct_op=pct(op_exp), pct_depr=pct(depr),
        pct_tax=pct(tax), pct_prive=pct(prive),
        pct_laba_kotor=pct(laba_kotor), pct_laba_op=pct(laba_op),
        pct_ebit=pct(ebit), pct_laba_bersih=pct(laba_bersih),
        pct_laba_tahan=pct(laba_tahan),
    )

def get_rekening_ids(conn):
    rows = conn.execute("SELECT id FROM akun WHERE is_rekening=1").fetchall()
    return [r['id'] for r in rows]

def get_rekening_saldo(conn):
    rows = conn.execute("""
        SELECT a.id, a.kode, a.nama, a.no_rekening,
               COALESCE(SUM(d.debit - d.kredit), 0) as saldo
        FROM akun a
        LEFT JOIN detail_jurnal d ON d.akun_id = a.id
        WHERE a.is_rekening = 1
        GROUP BY a.id ORDER BY a.kode
    """).fetchall()
    return [dict(r) for r in rows]

def calc_cashflow(conn, sd, ed):
    kas_ids = get_rekening_ids(conn)
    if not kas_ids:
        return {'masuk':0,'keluar':0,'saldo':0}
    ph = ','.join('?'*len(kas_ids))
    masuk = conn.execute(f"""
        SELECT COALESCE(SUM(d.debit),0) FROM detail_jurnal d
        JOIN jurnal j ON j.id=d.jurnal_id
        WHERE d.akun_id IN ({ph}) AND j.tanggal>=? AND j.tanggal<=?
    """, kas_ids+[sd,ed]).fetchone()[0]
    keluar = conn.execute(f"""
        SELECT COALESCE(SUM(d.kredit),0) FROM detail_jurnal d
        JOIN jurnal j ON j.id=d.jurnal_id
        WHERE d.akun_id IN ({ph}) AND j.tanggal>=? AND j.tanggal<=?
    """, kas_ids+[sd,ed]).fetchone()[0]
    saldo = masuk - keluar
    return {'masuk':masuk,'keluar':keluar,'saldo':saldo}

def month_end(y, m):
    last = calendar.monthrange(y, m)[1]
    return date(y, m, last)

def get_akun_id(conn, kode):
    r = conn.execute("SELECT id FROM akun WHERE kode=?", (kode,)).fetchone()
    return r['id'] if r else None

def insert_jurnal(conn, tanggal, keterangan, kategori, tipe_tx, entries):
    cur = conn.execute(
        "INSERT INTO jurnal(tanggal,keterangan,kategori,tipe_tx) VALUES(?,?,?,?)",
        (tanggal, keterangan, kategori, tipe_tx)
    )
    jid = cur.lastrowid
    for akun_kode, debit, kredit in entries:
        aid = get_akun_id(conn, akun_kode)
        if not aid:
            import logging
            logging.warning(f"insert_jurnal: akun kode '{akun_kode}' tidak ditemukan, entry dilewati (jurnal_id={jid})")
            continue
        if debit > 0 or kredit > 0:
            conn.execute(
                "INSERT INTO detail_jurnal(jurnal_id,akun_id,debit,kredit) VALUES(?,?,?,?)",
                (jid, aid, debit, kredit)
            )
    return jid

def update_piutang_status(conn, pid):
    p = conn.execute("SELECT jumlah,terbayar FROM piutang WHERE id=?", (pid,)).fetchone()
    if not p: return
    if p['terbayar'] >= p['jumlah']:
        s = 'LUNAS'
    elif p['terbayar'] > 0:
        s = 'SEBAGIAN'
    else:
        s = 'BELUM LUNAS'
    conn.execute("UPDATE piutang SET status=? WHERE id=?", (s, pid))

def update_hutang_status(conn, hid):
    h = conn.execute("SELECT jumlah,terbayar FROM hutang WHERE id=?", (hid,)).fetchone()
    if not h: return
    if h['terbayar'] >= h['jumlah']:
        s = 'LUNAS'
    elif h['terbayar'] > 0:
        s = 'SEBAGIAN'
    else:
        s = 'BELUM LUNAS'
    conn.execute("UPDATE hutang SET status=? WHERE id=?", (s, hid))

def piutang_status_label(row, today):
    if row['status'] == 'LUNAS':
        return 'LUNAS', 'success'
    sisa = row['jumlah'] - row['terbayar']
    if not row['jatuh_tempo']:
        return 'BELUM LUNAS', 'secondary'
    jt = datetime.strptime(str(row['jatuh_tempo'])[:10], '%Y-%m-%d').date()
    delta = (jt - today).days
    if delta < 0:
        return 'LEWAT JATUH TEMPO', 'danger'
    elif delta <= 7:
        return 'MENDEKATI JATUH TEMPO', 'warning'
    return 'BELUM JATUH TEMPO', 'info'

def get_setting(conn, key, default=''):
    r = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return r['value'] if r else default

def add_log(conn, aksi, detail='', kategori='LAINNYA'):
    uid   = session.get('user_id')
    uname = session.get('username', 'system')
    conn.execute(
        "INSERT INTO log_aktivitas(user_id, username, aksi, detail, kategori) VALUES(?,?,?,?,?)",
        (uid, uname, aksi, detail, kategori)
    )

def auto_penyusutan(conn):
    """Catat penyusutan yang belum dijurnal untuk semua aset aktif hingga bulan berjalan."""
    today = date.today()
    asets = conn.execute(
        "SELECT * FROM aset_tetap WHERE aktif=1 AND penyusutan_bulan > 0"
    ).fetchall()
    changed = False
    for aset in asets:
        try:
            tgl_beli = datetime.strptime(str(aset['tanggal_beli'])[:10], '%Y-%m-%d').date()
            months_elapsed = (today.year - tgl_beli.year) * 12 + (today.month - tgl_beli.month) + 1
            months_to_record = min(months_elapsed, aset['masa_pakai'])
            months_recorded  = aset['bulan_penyusutan_dicatat'] or 0
            if months_to_record <= months_recorded:
                continue
            months_new = months_to_record - months_recorded
            amount     = round(months_new * aset['penyusutan_bulan'], 2)
            if amount <= 0:
                continue
            ket = (f"Penyusutan {aset['nama']}: "
                   f"bulan {months_recorded+1}–{months_to_record} / {aset['masa_pakai']}")
            insert_jurnal(conn, str(today), ket, 'OPERASIONAL', 'PENYUSUTAN', [
                ('6130', amount, 0),
                ('1290', 0, amount),
            ])
            conn.execute(
                "UPDATE aset_tetap SET bulan_penyusutan_dicatat=?, akumulasi_penyusutan=akumulasi_penyusutan+? WHERE id=?",
                (months_to_record, amount, aset['id'])
            )
            changed = True
        except Exception:
            continue
    return changed


# ---------- DASHBOARD ----------
@app.route('/')
@login_required
def dashboard():
    today = date.today()
    sd = request.args.get('sd', today.replace(day=1).strftime('%Y-%m-%d'))
    ed = request.args.get('ed', today.strftime('%Y-%m-%d'))

    conn = db()
    if auto_penyusutan(conn):
        conn.commit()
    pnl   = calc_profitability(conn, sd, ed)
    cf    = calc_cashflow(conn, sd, ed)

    total_piutang = conn.execute(
        "SELECT COALESCE(SUM(jumlah-terbayar),0) FROM piutang WHERE status!='LUNAS'"
    ).fetchone()[0]
    total_hutang = conn.execute(
        "SELECT COALESCE(SUM(jumlah-terbayar),0) FROM hutang WHERE status!='LUNAS'"
    ).fetchone()[0]

    piutang_count = conn.execute(
        "SELECT COUNT(*) FROM piutang WHERE status!='LUNAS'"
    ).fetchone()[0]
    hutang_count = conn.execute(
        "SELECT COUNT(*) FROM hutang WHERE status!='LUNAS'"
    ).fetchone()[0]
    piutang_overdue = conn.execute(
        "SELECT COUNT(*) FROM piutang WHERE status!='LUNAS' AND jatuh_tempo < ?", (today,)
    ).fetchone()[0]
    hutang_overdue = conn.execute(
        "SELECT COUNT(*) FROM hutang WHERE status!='LUNAS' AND jatuh_tempo < ?", (today,)
    ).fetchone()[0]

    piutang_lunas = conn.execute(
        "SELECT COUNT(*) FROM piutang WHERE status='LUNAS'"
    ).fetchone()[0]
    piutang_total = conn.execute(
        "SELECT COUNT(*) FROM piutang"
    ).fetchone()[0]
    hutang_lunas = conn.execute(
        "SELECT COUNT(*) FROM hutang WHERE status='LUNAS'"
    ).fetchone()[0]
    hutang_total = conn.execute(
        "SELECT COUNT(*) FROM hutang"
    ).fetchone()[0]
    hutang_jatuh = conn.execute(
        "SELECT * FROM hutang WHERE status!='LUNAS' AND jatuh_tempo <= ? ORDER BY jatuh_tempo LIMIT 5",
        ((today + timedelta(days=7)),)
    ).fetchall()

    transaksi_baru = conn.execute("""
        SELECT j.id, j.tanggal, j.keterangan, j.referensi, j.kategori, j.tipe_tx,
               COALESCE(SUM(d.debit),0) as total
        FROM jurnal j LEFT JOIN detail_jurnal d ON d.jurnal_id=j.id
        GROUP BY j.id ORDER BY j.tanggal DESC, j.id DESC LIMIT 10
    """).fetchall()

    stok_rendah = conn.execute(
        "SELECT * FROM produk WHERE stok <= min_stok AND min_stok > 0 ORDER BY stok LIMIT 5"
    ).fetchall()
    piutang_jatuh = conn.execute(
        "SELECT * FROM piutang WHERE status!='LUNAS' AND jatuh_tempo <= ? ORDER BY jatuh_tempo LIMIT 5",
        ((today + timedelta(days=7)),)
    ).fetchall()

    nama_usaha = get_setting(conn, 'nama_usaha', 'Usaha Saya')
    saldo_rekening = get_rekening_saldo(conn)
    earliest_row = conn.execute("SELECT MIN(tanggal) FROM jurnal").fetchone()
    earliest_date = earliest_row[0] if earliest_row and earliest_row[0] else today.strftime('%Y-%m-%d')

    available_years = [row[0] for row in conn.execute(
        "SELECT DISTINCT CAST(strftime('%Y', tanggal) AS INTEGER) FROM jurnal ORDER BY 1 DESC"
    ).fetchall()]
    if not available_years:
        available_years = [today.year]

    # Mini insight: pendapatan tertinggi bulanan tahun ini
    year = today.year
    monthly_rev = []
    for m in range(1, today.month + 1):
        last_day = calendar.monthrange(year, m)[1]
        ms = f'{year}-{m:02d}-01'
        me = f'{year}-{m:02d}-{last_day:02d}'
        v = _qv(conn, "a.tipe='PENDAPATAN' AND j.tanggal>=? AND j.tanggal<=?", [ms, me])
        monthly_rev.append(v)
    max_rev_year = max(monthly_rev) if monthly_rev else 0

    # ── Breakdown cashflow: masuk per rekening, keluar per kategori ──────────
    cf_masuk_rek = [
        {'nama': r['nama'], 'total': r['total']}
        for r in conn.execute("""
            SELECT a.nama, COALESCE(SUM(d.debit),0) AS total
            FROM detail_jurnal d
            JOIN jurnal j ON j.id = d.jurnal_id
            JOIN akun a ON a.id = d.akun_id
            WHERE a.is_rekening=1 AND d.debit>0
              AND j.tanggal>=? AND j.tanggal<=?
            GROUP BY a.id, a.nama
            HAVING total>0 ORDER BY total DESC
        """, [sd, ed]).fetchall()
    ]
    cf_keluar_kat = [
        {'label': r['label'], 'total': r['total']}
        for r in conn.execute("""
            SELECT
              CASE
                WHEN a.kode LIKE '113%'               THEN 'Bahan Baku'
                WHEN a.kode LIKE '12%'                THEN 'Investasi Aset'
                WHEN a.kode = '6170'                  THEN 'Pajak'
                WHEN a.kode LIKE '61%'                THEN 'Beban Operasional'
                WHEN a.kode = '3300'                  THEN 'Prive'
                WHEN a.kode LIKE '2%'                 THEN 'Pelunasan Hutang'
                ELSE 'Lainnya'
              END AS label,
              COALESCE(SUM(d.debit),0) AS total
            FROM detail_jurnal d
            JOIN jurnal j ON j.id = d.jurnal_id
            JOIN akun a ON a.id = d.akun_id
            WHERE a.is_rekening=0 AND d.debit>0
              AND j.tanggal>=? AND j.tanggal<=?
              AND EXISTS (
                  SELECT 1 FROM detail_jurnal d2
                  JOIN akun a2 ON a2.id = d2.akun_id
                  WHERE d2.jurnal_id = d.jurnal_id
                    AND a2.is_rekening=1 AND d2.kredit>0
              )
            GROUP BY 1 HAVING total>0 ORDER BY total DESC
        """, [sd, ed]).fetchall()
    ]

    conn.close()

    session['dash_sd'] = sd
    session['dash_ed'] = ed

    return render_template('dashboard.html',
        sd=sd, ed=ed, pnl=pnl, cf=cf,
        max_rev_year=max_rev_year,
        total_piutang=total_piutang, total_hutang=total_hutang,
        piutang_count=piutang_count, hutang_count=hutang_count,
        piutang_overdue=piutang_overdue, hutang_overdue=hutang_overdue,
        piutang_lunas=piutang_lunas, piutang_total=piutang_total,
        hutang_lunas=hutang_lunas, hutang_total=hutang_total,
        hutang_jatuh=hutang_jatuh,
        transaksi_baru=transaksi_baru,
        stok_rendah=stok_rendah, piutang_jatuh=piutang_jatuh,
        nama_usaha=nama_usaha, saldo_rekening=saldo_rekening,
        earliest_date=earliest_date,
        available_years=available_years,
        cf_masuk_rek=cf_masuk_rek, cf_keluar_kat=cf_keluar_kat,
    )


# ---------- PEMASUKAN ----------
@app.route('/pemasukan', methods=['GET','POST'])
@operator_required
def pemasukan():
    conn = db()
    produk_list = conn.execute("SELECT * FROM produk ORDER BY nama").fetchall()
    akun_kas = conn.execute("SELECT * FROM akun WHERE is_rekening=1 ORDER BY kode").fetchall()

    if request.method == 'POST':
        mode = request.form.get('mode','umum')
        tanggal = request.form['tanggal']
        keterangan = request.form.get('keterangan','Penjualan')
        akun_kas_kode = request.form.get('akun_kas','1100')

        if not _operator_date_ok(tanggal):
            conn.close()
            flash('Operator hanya bisa input transaksi untuk bulan berjalan.', 'warning')
            return redirect(url_for('pemasukan'))

        if mode == 'umum':
            nominal = parse_rp(request.form.get('nominal','0'))
            diskon  = parse_rp(request.form.get('diskon','0'))
            ongkir  = parse_rp(request.form.get('ongkir','0'))
            biaya_lain = parse_rp(request.form.get('biaya_lain','0'))
            uang_masuk = parse_rp(request.form.get('uang_masuk','0'))
            hpp = parse_rp(request.form.get('hpp','0'))
            total = max(0, nominal - diskon + ongkir + biaya_lain)
            uang_masuk = min(uang_masuk, total)
            piutang_nm = max(0, total - uang_masuk)

            if hpp > 0:
                saldo_1130 = _qv(conn, "a.kode='1130'", [])
                if hpp > saldo_1130 + 0.01:
                    conn.close()
                    flash(f'HPP (Rp {hpp:,.0f}) melebihi saldo persediaan (Rp {saldo_1130:,.0f}). Periksa kembali.', 'danger')
                    return redirect(url_for('pemasukan'))

            if total <= 0:
                conn.close()
                flash('Nominal penjualan harus lebih dari 0.', 'danger')
                return redirect(url_for('pemasukan'))

            entries = []
            if uang_masuk > 0:
                entries.append((akun_kas_kode, uang_masuk, 0))
            if piutang_nm > 0:
                entries.append(('1120', piutang_nm, 0))
            entries.append(('4100', 0, total))
            if hpp > 0:
                entries += [('5100', hpp, 0), ('1130', 0, hpp)]

            insert_jurnal(conn, tanggal, keterangan, 'OPERASIONAL', 'PEMASUKAN', entries)

            if piutang_nm > 0:
                jt = request.form.get('jatuh_tempo') or None
                pelanggan = request.form.get('pelanggan', 'Customer')
                conn.execute(
                    "INSERT INTO piutang(tanggal,jatuh_tempo,pelanggan,keterangan,jumlah) VALUES(?,?,?,?,?)",
                    (tanggal, jt, pelanggan, keterangan, piutang_nm)
                )

        else:  # per produk — multi-item
            produk_ids   = request.form.getlist('produk_id[]')
            qtys         = request.form.getlist('qty[]')
            harga_juals  = request.form.getlist('harga_jual[]')
            diskon_items = request.form.getlist('diskon_item[]')
            ongkir     = parse_rp(request.form.get('ongkir','0') or '0')
            biaya_lain = parse_rp(request.form.get('biaya_lain','0') or '0')
            uang_masuk = parse_rp(request.form.get('uang_masuk','0') or '0')
            pelanggan  = request.form.get('pelanggan','Customer')
            jt = request.form.get('jatuh_tempo') or None

            items = []
            for i, pid_str in enumerate(produk_ids):
                if not pid_str: continue
                pid   = int(pid_str)
                qty   = float(qtys[i] if i < len(qtys) else 1) or 1
                harga = parse_rp(harga_juals[i] if i < len(harga_juals) else '0')
                dis   = parse_rp(diskon_items[i] if i < len(diskon_items) else '0')
                produk = conn.execute("SELECT * FROM produk WHERE id=?", (pid,)).fetchone()
                if not produk: continue
                subtotal  = qty * harga - dis
                hpp_item  = qty * produk['harga_beli']
                items.append({'produk': produk, 'qty': qty, 'harga': harga,
                              'subtotal': subtotal, 'hpp': hpp_item})

            total_sub = sum(it['subtotal'] for it in items)
            total_hpp = sum(it['hpp'] for it in items)
            grand_total = total_sub + ongkir + biaya_lain
            uang_masuk  = min(uang_masuk, grand_total)
            piutang_nm  = max(0, grand_total - uang_masuk)

            if not keterangan:
                names = [it['produk']['nama'] for it in items]
                keterangan = 'Penjualan ' + ', '.join(names[:3]) + ('+' if len(names) > 3 else '')

            if total_hpp > 0:
                saldo_1130 = _qv(conn, "a.kode='1130'", [])
                if total_hpp > saldo_1130 + 0.01:
                    conn.close()
                    flash(f'Total HPP produk (Rp {total_hpp:,.0f}) melebihi saldo persediaan (Rp {saldo_1130:,.0f}). Periksa harga beli produk.', 'danger')
                    return redirect(url_for('pemasukan'))

            if grand_total <= 0:
                conn.close()
                flash('Total penjualan harus lebih dari 0.', 'danger')
                return redirect(url_for('pemasukan'))

            entries = []
            if uang_masuk > 0:
                entries.append((akun_kas_kode, uang_masuk, 0))
            if piutang_nm > 0:
                entries.append(('1120', piutang_nm, 0))
            entries.append(('4100', 0, grand_total))
            if total_hpp > 0:
                entries += [('5100', total_hpp, 0), ('1130', 0, total_hpp)]

            insert_jurnal(conn, tanggal, keterangan, 'OPERASIONAL', 'PEMASUKAN', entries)

            for item in items:
                conn.execute("UPDATE produk SET stok=stok-? WHERE id=?",
                             (item['qty'], item['produk']['id']))
                conn.execute(
                    "INSERT INTO pergerakan_stok(produk_id,tanggal,jenis,qty,harga,keterangan) VALUES(?,?,?,?,?,?)",
                    (item['produk']['id'], tanggal, 'KELUAR', item['qty'], item['harga'], keterangan)
                )

            if piutang_nm > 0:
                conn.execute(
                    "INSERT INTO piutang(tanggal,jatuh_tempo,pelanggan,keterangan,jumlah) VALUES(?,?,?,?,?)",
                    (tanggal, jt, pelanggan, keterangan, piutang_nm)
                )

        buat_invoice = request.form.get('buat_invoice') == 'on'
        inv_id = None
        if buat_invoice:
            inv_s   = _inv_settings(conn)
            inv_nomor = request.form.get('inv_nomor','').strip() or _next_inv_number(conn, inv_s['inv_prefix'])
            inv_tel   = request.form.get('inv_telepon_pelanggan','')
            inv_adr   = request.form.get('inv_alamat_pelanggan','')
            inv_cat   = request.form.get('inv_catatan_custom','')
            # derive pelanggan & jatuh_tempo — invoice-specific fields take priority
            inv_pel   = request.form.get('inv_pelanggan','').strip() or request.form.get('pelanggan','').strip() or 'Customer'
            inv_jt    = request.form.get('inv_jatuh_tempo','').strip() or request.form.get('jatuh_tempo','').strip() or None
            # derive total for invoice
            if mode == 'umum':
                nominal_u  = parse_rp(request.form.get('nominal','0'))
                diskon_u   = parse_rp(request.form.get('diskon','0'))
                ongkir_u   = parse_rp(request.form.get('ongkir','0'))
                biaya_u    = parse_rp(request.form.get('biaya_lain','0'))
                inv_total  = max(0, nominal_u - diskon_u + ongkir_u + biaya_u)
                inv_diskon = diskon_u; inv_ongkir = ongkir_u; inv_biaya = biaya_u
                inv_items  = [{'deskripsi': keterangan or 'Penjualan', 'qty': 1, 'satuan': 'unit',
                                'harga_satuan': nominal_u, 'diskon_item': 0,
                                'subtotal': nominal_u}]
            else:  # produk
                inv_ongkir  = parse_rp(request.form.get('ongkir','0') or '0')
                inv_biaya   = parse_rp(request.form.get('biaya_lain','0') or '0')
                inv_diskon  = 0
                inv_items   = []
                _pids  = request.form.getlist('produk_id[]')
                _qtys  = request.form.getlist('qty[]')
                _hjs   = request.form.getlist('harga_jual[]')
                _diss  = request.form.getlist('diskon_item[]')
                for i, pid_str in enumerate(_pids):
                    if not pid_str: continue
                    p_row = conn.execute("SELECT * FROM produk WHERE id=?", (int(pid_str),)).fetchone()
                    if not p_row: continue
                    qty_i = float(_qtys[i] if i < len(_qtys) else 1) or 1
                    hj_i  = parse_rp(_hjs[i] if i < len(_hjs) else '0')
                    dis_i = parse_rp(_diss[i] if i < len(_diss) else '0')
                    inv_items.append({'deskripsi': p_row['nama'], 'qty': qty_i,
                                      'satuan': p_row['satuan'] or 'pcs',
                                      'harga_satuan': hj_i, 'diskon_item': dis_i,
                                      'subtotal': qty_i * hj_i - dis_i})
            conn.execute(
                "INSERT INTO invoice(nomor,tanggal,jatuh_tempo,pelanggan,alamat_pelanggan,"
                "telepon_pelanggan,diskon,ongkir,biaya_lain,catatan,status,dibuat) VALUES(?,?,?,?,?,?,?,?,?,?,'DRAFT',?)",
                (inv_nomor, tanggal, inv_jt, inv_pel, inv_adr, inv_tel,
                 inv_diskon, inv_ongkir, inv_biaya, inv_cat, tanggal)
            )
            inv_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            for it in inv_items:
                conn.execute(
                    "INSERT INTO invoice_item(invoice_id,deskripsi,qty,satuan,harga_satuan,diskon_item,subtotal) VALUES(?,?,?,?,?,?,?)",
                    (inv_id, it['deskripsi'], it['qty'], it['satuan'],
                     it['harga_satuan'], it['diskon_item'], it['subtotal'])
                )

        # Tag nomor invoice ke keterangan piutang agar lookup di cetak invoice akurat
        if inv_id and inv_nomor:
            conn.execute(
                """UPDATE piutang SET keterangan = keterangan || ' | ' || ?
                   WHERE id = (
                       SELECT id FROM piutang WHERE pelanggan=? AND tanggal=?
                       ORDER BY id DESC LIMIT 1
                   )""",
                (inv_nomor, request.form.get('pelanggan', request.form.get('inv_pelanggan', '')), tanggal)
            )

        add_log(conn, 'Pemasukan dicatat', f"{keterangan} | {tanggal}", 'INPUT')
        conn.commit(); conn.close()
        flash('Pemasukan berhasil dicatat!', 'success')
        if inv_id:
            return redirect(url_for('pemasukan', new_inv=inv_id))
        return redirect(url_for('pemasukan'))

    saldo_persediaan = _qv(conn, "a.kode='1130'", [])
    conn.close()
    return render_template('pemasukan.html', produk_list=produk_list, akun_kas=akun_kas,
                           today=date.today().strftime('%Y-%m-%d'),
                           saldo_persediaan=saldo_persediaan)

@app.route('/api/produk/<int:pid>')
@operator_required
def api_produk(pid):
    conn = db()
    p = conn.execute("SELECT * FROM produk WHERE id=?", (pid,)).fetchone()
    conn.close()
    if not p:
        return jsonify({})
    return jsonify({'harga_beli': p['harga_beli'], 'harga_jual': p['harga_jual'],
                    'stok': p['stok'], 'satuan': p['satuan']})


# ---------- PENGELUARAN ----------
OPERASIONAL_MAP = {
    'Gaji':'6100','Sewa':'6110','Utilitas':'6120','Pemasaran':'6140',
    'Administrasi':'6150','Bunga':'6160','Lainnya':'6180'
}

@app.route('/pengeluaran', methods=['GET','POST'])
@operator_required
def pengeluaran():
    conn = db()
    akun_kas = conn.execute("SELECT * FROM akun WHERE is_rekening=1 ORDER BY kode").fetchall()
    aset_list = conn.execute("SELECT * FROM aset_tetap WHERE aktif=1 ORDER BY nama").fetchall()

    if request.method == 'POST':
        kategori   = request.form['kategori']
        tanggal    = request.form['tanggal']
        keterangan = request.form.get('keterangan','')
        nominal    = parse_rp(request.form.get('nominal','0'))
        kas_mode   = request.form.get('kas_mode','single')

        if not _operator_date_ok(tanggal):
            conn.close()
            flash('Operator hanya bisa input transaksi untuk bulan berjalan.', 'warning')
            return redirect(url_for('pengeluaran'))

        # ── Hutang fields (semua kategori sekarang support hutang) ───────────────
        pemasok_raw = (request.form.get('pemasok','') or '').strip()
        jt          = request.form.get('jatuh_tempo') or None

        # ── Build kas credit entries + tentukan uang_keluar ──────────────────────
        if kas_mode == 'multi':
            kas_kodens = request.form.getlist('akun_kas_multi[]')
            kas_noms   = request.form.getlist('nominal_kas_multi[]')
            kas_splits = [(k, float(n or 0)) for k, n in zip(kas_kodens, kas_noms)
                          if k and float(n or 0) > 0]
            total_kas  = sum(n for _, n in kas_splits)
            if total_kas > nominal:
                conn.close()
                flash('Total pembayaran multi-rekening melebihi nominal transaksi. Periksa kembali.', 'danger')
                return redirect(url_for('pengeluaran'))
            uang_keluar = total_kas
            def kas_credits(amount):
                return [(kode, 0, nom) for kode, nom in kas_splits]
        else:
            akun_kas_kode = request.form.get('akun_kas','1100')
            raw_uk = request.form.get('uang_keluar')
            if raw_uk is None:
                # Field tidak ada di form (form versi lama) → default full cash semua kategori
                uang_keluar = nominal
            else:
                # Field ada (mungkin kosong) → kosong berarti user memang ingin uang_keluar=0
                uang_keluar = parse_rp(str(raw_uk))
            def kas_credits(amount):
                return [(akun_kas_kode, 0, amount)] if amount > 0 else []

        uang_keluar = max(0, min(uang_keluar, nominal))
        hutang_nm   = nominal - uang_keluar

        if kategori == 'OPERASIONAL':
            sub = request.form.get('subkategori','Lainnya')
            beban_kode = OPERASIONAL_MAP.get(sub, '6180')
            keterangan = keterangan or f"Beban {sub}"
            entries = [(beban_kode, nominal, 0)] + kas_credits(uang_keluar)
            if hutang_nm > 0: entries.append(('2100', 0, hutang_nm))
            insert_jurnal(conn, tanggal, keterangan, 'OPERASIONAL', 'PENGELUARAN', entries)

        elif kategori == 'PAJAK':
            keterangan = keterangan or 'Pembayaran Pajak'
            entries = [('6170', nominal, 0)] + kas_credits(uang_keluar)
            if hutang_nm > 0: entries.append(('2100', 0, hutang_nm))
            insert_jurnal(conn, tanggal, keterangan, 'OPERASIONAL', 'PENGELUARAN', entries)

        elif kategori == 'PENARIKAN_OWNER':
            keterangan = keterangan or 'Penarikan Owner / Prive'
            entries = [('3300', nominal, 0)] + kas_credits(uang_keluar)
            if hutang_nm > 0: entries.append(('2100', 0, hutang_nm))
            insert_jurnal(conn, tanggal, keterangan, 'PENDANAAN', 'PENGELUARAN', entries)

        elif kategori in ('BAHAN_BAKU', 'INVESTASI_ASET'):
            if kategori == 'BAHAN_BAKU':
                keterangan = keterangan or 'Belanja Bahan Baku'
                entries = [('1130', nominal, 0)] + kas_credits(uang_keluar)
                if hutang_nm > 0: entries.append(('2100', 0, hutang_nm))
                insert_jurnal(conn, tanggal, keterangan, 'OPERASIONAL', 'PENGELUARAN', entries)
                # Opsional: buat produk baru atau link ke produk yang sudah ada
                buat_produk_baru = request.form.get('buat_produk_baru') == '1'
                nama_produk_baru = request.form.get('nama_produk_baru', '').strip()
                # Qty dibaca dari field sesuai mode (2 field beda nama agar tidak bentrok)
                _raw_qty = request.form.get('qty_bb_baru') if buat_produk_baru else request.form.get('qty_bb')
                try:
                    qty_bb = float(_raw_qty or 0)
                except (ValueError, TypeError):
                    qty_bb = 0.0

                if buat_produk_baru and nama_produk_baru:
                    import re as _re
                    kode_base = _re.sub(r'[^A-Z0-9]', '-', nama_produk_baru.upper())[:14].strip('-')
                    kode_new = kode_base; _sfx = 1
                    while conn.execute('SELECT id FROM produk WHERE kode=?', (kode_new,)).fetchone():
                        kode_new = f'{kode_base}-{_sfx}'; _sfx += 1
                    satuan_baru     = request.form.get('satuan_produk_baru', 'pcs').strip() or 'pcs'
                    harga_jual_baru = parse_rp(request.form.get('harga_jual_produk_baru', '0'))
                    hpp_input       = request.form.get('harga_beli_produk_baru', '').strip()
                    harga_beli_baru = parse_rp(hpp_input) if hpp_input else (
                                      round(nominal / qty_bb, 2) if qty_bb > 0 else nominal)
                    min_stok_baru   = float(request.form.get('min_stok_produk_baru', 0) or 0)
                    conn.execute(
                        'INSERT INTO produk(kode,nama,satuan,harga_beli,harga_jual,stok,min_stok) VALUES(?,?,?,?,?,0,?)',
                        (kode_new, nama_produk_baru, satuan_baru, harga_beli_baru, harga_jual_baru, min_stok_baru)
                    )
                    produk_id_bb = str(conn.execute('SELECT last_insert_rowid()').fetchone()[0])
                else:
                    produk_id_bb = request.form.get('produk_id_bb')

                if produk_id_bb and qty_bb > 0:
                    conn.execute('UPDATE produk SET stok=stok+? WHERE id=?', (qty_bb, int(produk_id_bb)))
                    conn.execute(
                        'INSERT INTO pergerakan_stok(produk_id,tanggal,jenis,qty,harga,keterangan) VALUES(?,?,?,?,?,?)',
                        (int(produk_id_bb), tanggal, 'MASUK', qty_bb,
                         round(nominal / qty_bb, 2) if qty_bb > 0 else 0, keterangan)
                    )
            else:
                nama_aset  = request.form.get('nama_aset','Aset Baru')
                kat_aset   = request.form.get('kategori_aset','Peralatan')
                masa_pakai = int(request.form.get('masa_pakai',12) or 12)
                peny_bln   = round(nominal / masa_pakai, 2) if masa_pakai > 0 else 0
                aset_kode_map = {'Peralatan':'1200','Kendaraan':'1210','Gedung':'1220'}
                aset_kode  = aset_kode_map.get(kat_aset,'1200')
                keterangan = keterangan or f"Investasi {kat_aset}: {nama_aset}"
                entries = [(aset_kode, nominal, 0)] + kas_credits(uang_keluar)
                if hutang_nm > 0: entries.append(('2100', 0, hutang_nm))
                insert_jurnal(conn, tanggal, keterangan, 'INVESTASI', 'PENGELUARAN', entries)
                conn.execute(
                    "INSERT INTO aset_tetap(nama,kategori,harga_beli,tanggal_beli,masa_pakai,penyusutan_bulan) VALUES(?,?,?,?,?,?)",
                    (nama_aset, kat_aset, nominal, tanggal, masa_pakai, peny_bln)
                )

        # ── Insert ke tabel hutang (berlaku untuk SEMUA kategori sekarang) ───────
        if hutang_nm > 0:
            _default_pemasok = {
                'BAHAN_BAKU':       'Pemasok',
                'INVESTASI_ASET':   'Vendor Aset',
                'OPERASIONAL':      'Kreditur Operasional',
                'PAJAK':            'Kantor Pajak',
                'PENARIKAN_OWNER':  'Hutang ke Owner',
            }.get(kategori, 'Kreditur')
            nama_kreditur = pemasok_raw or _default_pemasok
            conn.execute(
                "INSERT INTO hutang(tanggal,jatuh_tempo,pemasok,keterangan,jumlah) VALUES(?,?,?,?,?)",
                (tanggal, jt, nama_kreditur, keterangan, hutang_nm)
            )

        add_log(conn, 'Pengeluaran dicatat', f"{keterangan} | {tanggal} | {kategori}", 'INPUT')
        conn.commit(); conn.close()
        flash('Pengeluaran berhasil dicatat!', 'success')
        return redirect(url_for('dashboard'))

    produk_list_bb = conn.execute("SELECT id, nama, satuan, stok FROM produk ORDER BY nama").fetchall()
    conn.close()
    return render_template('pengeluaran.html', akun_kas=akun_kas, aset_list=aset_list,
                           today=date.today().strftime('%Y-%m-%d'),
                           subkategori_list=list(OPERASIONAL_MAP.keys()),
                           produk_list_bb=produk_list_bb)


# ---------- RETUR PENJUALAN ----------
@app.route('/retur-penjualan', methods=['GET','POST'])
@operator_required
def retur_penjualan():
    conn = db()
    produk_list = conn.execute("SELECT * FROM produk ORDER BY nama").fetchall()
    akun_kas = conn.execute("SELECT * FROM akun WHERE is_rekening=1 ORDER BY kode").fetchall()

    if request.method == 'POST':
        mode       = request.form.get('mode', 'umum')
        tanggal    = request.form['tanggal']
        keterangan = request.form.get('keterangan', 'Retur Penjualan')
        sumber     = request.form.get('sumber', 'kas')   # 'kas' atau 'piutang'
        akun_kas_kode = request.form.get('akun_kas', '1100')

        if not _operator_date_ok(tanggal):
            conn.close()
            flash('Operator hanya bisa input transaksi untuk bulan berjalan.', 'warning')
            return redirect(url_for('retur_penjualan'))

        if mode == 'umum':
            nominal = parse_rp(request.form.get('nominal', '0'))
            hpp_balik = parse_rp(request.form.get('hpp_balik', '0'))
            if nominal <= 0:
                conn.close()
                flash('Nominal retur harus lebih dari 0.', 'danger')
                return redirect(url_for('retur_penjualan'))

            entries = [('4150', nominal, 0)]   # debit kontra-pendapatan
            if sumber == 'piutang':
                entries.append(('1120', 0, nominal))   # kurangi piutang
            else:
                entries.append((akun_kas_kode, 0, nominal))   # uang keluar
            # HPP reversal opsional
            if hpp_balik > 0:
                entries += [('1130', hpp_balik, 0), ('5100', 0, hpp_balik)]

            insert_jurnal(conn, tanggal, keterangan, 'OPERASIONAL', 'RETUR_JUAL', entries)

            # Kurangi piutang record jika sumber=piutang
            if sumber == 'piutang':
                pelanggan = request.form.get('pelanggan', '').strip()
                if pelanggan:
                    p = conn.execute(
                        "SELECT id, jumlah FROM piutang WHERE pelanggan=? AND status!='LUNAS' ORDER BY tanggal DESC LIMIT 1",
                        (pelanggan,)
                    ).fetchone()
                    if p:
                        new_jml = max(0, p['jumlah'] - nominal)
                        conn.execute("UPDATE piutang SET jumlah=? WHERE id=?", (new_jml, p['id']))
                        update_piutang_status(conn, p['id'])

        else:  # per produk
            produk_ids   = request.form.getlist('produk_id[]')
            qtys         = request.form.getlist('qty[]')
            harga_juals  = request.form.getlist('harga_jual[]')

            items = []
            for i, pid_str in enumerate(produk_ids):
                if not pid_str: continue
                pid   = int(pid_str)
                qty   = float(qtys[i] if i < len(qtys) else 1) or 1
                harga = parse_rp(harga_juals[i] if i < len(harga_juals) else '0')
                produk = conn.execute("SELECT * FROM produk WHERE id=?", (pid,)).fetchone()
                if not produk: continue
                subtotal = qty * harga
                hpp_item = qty * produk['harga_beli']
                items.append({'produk': produk, 'qty': qty, 'harga': harga,
                              'subtotal': subtotal, 'hpp': hpp_item})

            total_sub = sum(it['subtotal'] for it in items)
            total_hpp = sum(it['hpp'] for it in items)

            if total_sub <= 0 or not items:
                conn.close()
                flash('Pilih minimal 1 produk dengan nominal > 0.', 'danger')
                return redirect(url_for('retur_penjualan'))

            entries = [('4150', total_sub, 0)]
            if sumber == 'piutang':
                entries.append(('1120', 0, total_sub))
            else:
                entries.append((akun_kas_kode, 0, total_sub))
            if total_hpp > 0:
                entries += [('1130', total_hpp, 0), ('5100', 0, total_hpp)]

            insert_jurnal(conn, tanggal, keterangan, 'OPERASIONAL', 'RETUR_JUAL', entries)

            # Stok masuk kembali
            for item in items:
                conn.execute("UPDATE produk SET stok=stok+? WHERE id=?",
                             (item['qty'], item['produk']['id']))
                conn.execute(
                    "INSERT INTO pergerakan_stok(produk_id,tanggal,jenis,qty,harga,keterangan) VALUES(?,?,?,?,?,?)",
                    (item['produk']['id'], tanggal, 'MASUK', item['qty'], item['harga'],
                     f"Retur: {keterangan}")
                )

            if sumber == 'piutang':
                pelanggan = request.form.get('pelanggan', '').strip()
                if pelanggan:
                    p = conn.execute(
                        "SELECT id, jumlah FROM piutang WHERE pelanggan=? AND status!='LUNAS' ORDER BY tanggal DESC LIMIT 1",
                        (pelanggan,)
                    ).fetchone()
                    if p:
                        new_jml = max(0, p['jumlah'] - total_sub)
                        conn.execute("UPDATE piutang SET jumlah=? WHERE id=?", (new_jml, p['id']))
                        update_piutang_status(conn, p['id'])

        add_log(conn, 'Retur Penjualan', f"{keterangan} | {tanggal}", 'INPUT')
        conn.commit(); conn.close()
        flash('Retur penjualan berhasil dicatat!', 'success')
        return redirect(url_for('retur_penjualan'))

    saldo_persediaan = _qv(conn, "a.kode='1130'", [])
    conn.close()
    return render_template('retur_penjualan.html',
                           produk_list=produk_list, akun_kas=akun_kas,
                           today=date.today().strftime('%Y-%m-%d'),
                           saldo_persediaan=saldo_persediaan)


# ---------- RETUR PEMBELIAN ----------
@app.route('/retur-pembelian', methods=['GET','POST'])
@operator_required
def retur_pembelian():
    conn = db()
    produk_list = conn.execute("SELECT * FROM produk ORDER BY nama").fetchall()
    akun_kas = conn.execute("SELECT * FROM akun WHERE is_rekening=1 ORDER BY kode").fetchall()

    if request.method == 'POST':
        mode       = request.form.get('mode', 'umum')
        tanggal    = request.form['tanggal']
        keterangan = request.form.get('keterangan', 'Retur Pembelian')
        sumber     = request.form.get('sumber', 'kas')   # 'kas' atau 'hutang'
        akun_kas_kode = request.form.get('akun_kas', '1100')

        if not _operator_date_ok(tanggal):
            conn.close()
            flash('Operator hanya bisa input transaksi untuk bulan berjalan.', 'warning')
            return redirect(url_for('retur_pembelian'))

        if mode == 'umum':
            nominal = parse_rp(request.form.get('nominal', '0'))
            if nominal <= 0:
                conn.close()
                flash('Nominal retur harus lebih dari 0.', 'danger')
                return redirect(url_for('retur_pembelian'))

            # App ini pakai perpetual inventory: pembelian langsung ke 1130, jadi
            # retur langsung mengurangi 1130. Akun 5150 tersedia di COA untuk
            # entry manual periodic-style bila diperlukan.
            if sumber == 'hutang':
                entries = [('2100', nominal, 0), ('1130', 0, nominal)]
            else:
                entries = [(akun_kas_kode, nominal, 0), ('1130', 0, nominal)]

            insert_jurnal(conn, tanggal, keterangan, 'OPERASIONAL', 'RETUR_BELI', entries)

            if sumber == 'hutang':
                pemasok = request.form.get('pemasok', '').strip()
                if pemasok:
                    h = conn.execute(
                        "SELECT id, jumlah FROM hutang WHERE pemasok=? AND status!='LUNAS' ORDER BY tanggal DESC LIMIT 1",
                        (pemasok,)
                    ).fetchone()
                    if h:
                        new_jml = max(0, h['jumlah'] - nominal)
                        conn.execute("UPDATE hutang SET jumlah=? WHERE id=?", (new_jml, h['id']))
                        update_hutang_status(conn, h['id'])

        else:  # per produk
            produk_ids = request.form.getlist('produk_id[]')
            qtys       = request.form.getlist('qty[]')
            hargas     = request.form.getlist('harga_beli[]')

            items = []
            for i, pid_str in enumerate(produk_ids):
                if not pid_str: continue
                pid   = int(pid_str)
                qty   = float(qtys[i] if i < len(qtys) else 1) or 1
                harga = parse_rp(hargas[i] if i < len(hargas) else '0')
                produk = conn.execute("SELECT * FROM produk WHERE id=?", (pid,)).fetchone()
                if not produk: continue
                if qty > produk['stok']:
                    conn.close()
                    flash(f'Stok {produk["nama"]} tidak cukup untuk retur ({produk["stok"]} < {qty}).', 'danger')
                    return redirect(url_for('retur_pembelian'))
                items.append({'produk': produk, 'qty': qty, 'harga': harga,
                              'subtotal': qty * harga})

            total_sub = sum(it['subtotal'] for it in items)
            if total_sub <= 0 or not items:
                conn.close()
                flash('Pilih minimal 1 produk dengan nominal > 0.', 'danger')
                return redirect(url_for('retur_pembelian'))

            if sumber == 'hutang':
                entries = [('2100', total_sub, 0), ('1130', 0, total_sub)]
            else:
                entries = [(akun_kas_kode, total_sub, 0), ('1130', 0, total_sub)]

            insert_jurnal(conn, tanggal, keterangan, 'OPERASIONAL', 'RETUR_BELI', entries)

            # Stok keluar (kembali ke supplier)
            for item in items:
                conn.execute("UPDATE produk SET stok=stok-? WHERE id=?",
                             (item['qty'], item['produk']['id']))
                conn.execute(
                    "INSERT INTO pergerakan_stok(produk_id,tanggal,jenis,qty,harga,keterangan) VALUES(?,?,?,?,?,?)",
                    (item['produk']['id'], tanggal, 'KELUAR', item['qty'], item['harga'],
                     f"Retur ke supplier: {keterangan}")
                )

            if sumber == 'hutang':
                pemasok = request.form.get('pemasok', '').strip()
                if pemasok:
                    h = conn.execute(
                        "SELECT id, jumlah FROM hutang WHERE pemasok=? AND status!='LUNAS' ORDER BY tanggal DESC LIMIT 1",
                        (pemasok,)
                    ).fetchone()
                    if h:
                        new_jml = max(0, h['jumlah'] - total_sub)
                        conn.execute("UPDATE hutang SET jumlah=? WHERE id=?", (new_jml, h['id']))
                        update_hutang_status(conn, h['id'])

        add_log(conn, 'Retur Pembelian', f"{keterangan} | {tanggal}", 'INPUT')
        conn.commit(); conn.close()
        flash('Retur pembelian berhasil dicatat!', 'success')
        return redirect(url_for('retur_pembelian'))

    saldo_persediaan = _qv(conn, "a.kode='1130'", [])
    conn.close()
    return render_template('retur_pembelian.html',
                           produk_list=produk_list, akun_kas=akun_kas,
                           today=date.today().strftime('%Y-%m-%d'),
                           saldo_persediaan=saldo_persediaan)


# ---------- PELUNASAN ----------
@app.route('/pelunasan', methods=['GET','POST'])
@finance_required
def pelunasan():
    conn = db()
    piutang_aktif = conn.execute(
        "SELECT * FROM piutang WHERE status!='LUNAS' ORDER BY jatuh_tempo, tanggal"
    ).fetchall()
    hutang_aktif = conn.execute(
        "SELECT * FROM hutang WHERE status!='LUNAS' ORDER BY jatuh_tempo, tanggal"
    ).fetchall()
    akun_kas = conn.execute("SELECT * FROM akun WHERE is_rekening=1 ORDER BY kode").fetchall()

    if request.method == 'POST':
        jenis   = request.form['jenis']
        # getlist karena ada 2 select dengan name=record_id (piutang + hutang)
        rid_candidates = [v.strip() for v in request.form.getlist('record_id') if v.strip()]
        if not rid_candidates:
            conn.close()
            flash('Pilih hutang atau piutang terlebih dahulu.', 'warning')
            return redirect(url_for('pelunasan'))
        rid_str = rid_candidates[-1]  # ambil yang terakhir (select yang visible)
        rid    = int(rid_str)
        nominal= float(request.form.get('nominal',0) or 0)
        tanggal= request.form.get('tanggal', str(date.today()))
        akun_kas_kode = request.form.get('akun_kas','1100')
        catatan= request.form.get('catatan','')

        if jenis == 'PIUTANG':
            p = conn.execute("SELECT * FROM piutang WHERE id=?", (rid,)).fetchone()
            if p:
                sisa = p['jumlah'] - p['terbayar']
                nominal = min(nominal, sisa)
                if nominal <= 0:
                    conn.close(); flash('Piutang sudah lunas atau jumlah tidak valid.', 'warning')
                    return redirect(url_for('pelunasan'))
                conn.execute("INSERT INTO bayar_piutang(piutang_id,tanggal,jumlah,catatan) VALUES(?,?,?,?)",
                             (rid, tanggal, nominal, catatan))
                conn.execute("UPDATE piutang SET terbayar=terbayar+? WHERE id=?", (nominal, rid))
                update_piutang_status(conn, rid)
                insert_jurnal(conn, tanggal, f"Pelunasan piutang: {p['pelanggan']}", 'OPERASIONAL', 'PELUNASAN', [
                    (akun_kas_kode, nominal, 0), ('1120', 0, nominal)
                ])
                add_log(conn, 'Pelunasan piutang', f"{p['pelanggan']} | Rp {nominal:,.0f} | {tanggal}", 'INPUT')
                flash('Pelunasan piutang dicatat!', 'success')
        else:
            h = conn.execute("SELECT * FROM hutang WHERE id=?", (rid,)).fetchone()
            if h:
                sisa = h['jumlah'] - h['terbayar']
                nominal = min(nominal, sisa)
                if nominal <= 0:
                    conn.close(); flash('Hutang sudah lunas atau jumlah tidak valid.', 'warning')
                    return redirect(url_for('pelunasan'))
                conn.execute("INSERT INTO bayar_hutang(hutang_id,tanggal,jumlah,catatan) VALUES(?,?,?,?)",
                             (rid, tanggal, nominal, catatan))
                conn.execute("UPDATE hutang SET terbayar=terbayar+? WHERE id=?", (nominal, rid))
                update_hutang_status(conn, rid)
                insert_jurnal(conn, tanggal, f"Pelunasan hutang: {h['pemasok']}", 'OPERASIONAL', 'PELUNASAN', [
                    ('2100', nominal, 0), (akun_kas_kode, 0, nominal)
                ])
                add_log(conn, 'Pelunasan hutang', f"{h['pemasok']} | Rp {nominal:,.0f} | {tanggal}", 'INPUT')
                flash('Pelunasan hutang dicatat!', 'success')

        conn.commit(); conn.close()
        return redirect(url_for('pelunasan'))

    conn.close()
    return render_template('pelunasan.html',
                           piutang_aktif=piutang_aktif, hutang_aktif=hutang_aktif,
                           akun_kas=akun_kas, today=date.today().strftime('%Y-%m-%d'))


# ---------- PIUTANG ----------
@app.route('/piutang')
@investor_required
def piutang_list():
    conn = db()
    today = date.today()
    sf = request.args.get('status','')
    q = request.args.get('q','')
    where, params = [], []
    if sf: where.append("status=?"); params.append(sf)
    if q:  where.append("(pelanggan LIKE ? OR keterangan LIKE ?)"); params += [f'%{q}%',f'%{q}%']
    ws = ('WHERE '+' AND '.join(where)) if where else ''
    rows = conn.execute(f'SELECT * FROM piutang {ws} ORDER BY CASE status WHEN "LUNAS" THEN 1 ELSE 0 END, jatuh_tempo, tanggal DESC', params).fetchall()
    total = conn.execute("SELECT COALESCE(SUM(jumlah-terbayar),0) FROM piutang WHERE status!='LUNAS'").fetchone()[0]
    rows_with_status = []
    for r in rows:
        lbl, cls = piutang_status_label(r, today)
        rows_with_status.append((r, lbl, cls))
    akun_kas = conn.execute("SELECT kode,nama FROM akun WHERE is_rekening=1 ORDER BY kode").fetchall()
    conn.close()
    return render_template('piutang.html', rows=rows_with_status, total=total, sf=sf, q=q,
                           akun_kas=akun_kas, today=date.today())

@app.route('/piutang/baru', methods=['POST'])
@operator_required
def piutang_baru():
    flash('Piutang tidak bisa ditambah langsung. Catat melalui form Pemasukan agar jurnal terbentuk.', 'warning')
    return redirect(url_for('piutang_list'))

@app.route('/piutang/<int:id>/edit', methods=['GET','POST'])
@operator_required
def piutang_edit(id):
    conn = db()
    p = conn.execute('SELECT * FROM piutang WHERE id=?', (id,)).fetchone()
    if not p:
        conn.close(); flash('Tidak ditemukan.','danger')
        return redirect(url_for('piutang_list'))
    if request.method == 'POST':
        conn.execute("""UPDATE piutang SET tanggal=?,jatuh_tempo=?,pelanggan=?,keterangan=? WHERE id=?""",
            (request.form['tanggal'], request.form.get('jatuh_tempo') or None,
             request.form['pelanggan'], request.form.get('keterangan',''), id))
        conn.commit(); conn.close()
        flash('Piutang diperbarui!', 'success')
        return redirect(url_for('piutang_list'))
    conn.close()
    return render_template('piutang_edit.html', p=p)

@app.route('/piutang/<int:id>/bayar', methods=['POST'])
@operator_required
def piutang_bayar(id):
    conn = db()
    jumlah = parse_rp(request.form.get('jumlah','0'))
    tanggal = request.form.get('tanggal', str(date.today()))
    catatan = request.form.get('catatan','')
    akun_kas_kode = request.form.get('akun_kas','1100')
    if not _operator_date_ok(tanggal):
        conn.close()
        flash('Operator hanya bisa mencatat pembayaran untuk bulan berjalan.', 'warning')
        return redirect(url_for('piutang_list'))
    p = conn.execute('SELECT * FROM piutang WHERE id=?', (id,)).fetchone()
    if not p:
        conn.close(); flash('Piutang tidak ditemukan.', 'danger')
        return redirect(url_for('piutang_list'))
    sisa = p['jumlah'] - p['terbayar']
    jumlah = min(jumlah, sisa)
    if jumlah <= 0:
        conn.close(); flash('Piutang sudah lunas atau jumlah tidak valid.', 'warning')
        return redirect(url_for('piutang_list'))
    conn.execute('INSERT INTO bayar_piutang(piutang_id,tanggal,jumlah,catatan) VALUES(?,?,?,?)',
                 (id, tanggal, jumlah, catatan))
    conn.execute('UPDATE piutang SET terbayar=terbayar+? WHERE id=?', (jumlah, id))
    update_piutang_status(conn, id)
    insert_jurnal(conn, tanggal, f"Pelunasan piutang: {p['pelanggan']}", 'OPERASIONAL', 'PELUNASAN', [
        (akun_kas_kode, jumlah, 0), ('1120', 0, jumlah)
    ])
    conn.commit(); conn.close()
    flash('Pembayaran piutang dicatat!', 'success')
    return redirect(url_for('piutang_list'))

@app.route('/piutang/<int:id>/riwayat')
@investor_required
def piutang_riwayat(id):
    conn = db()
    p = conn.execute('SELECT * FROM piutang WHERE id=?', (id,)).fetchone()
    if not p:
        conn.close(); flash('Piutang tidak ditemukan.', 'danger')
        return redirect(url_for('piutang_list'))
    riwayat = conn.execute('SELECT * FROM bayar_piutang WHERE piutang_id=? ORDER BY tanggal DESC', (id,)).fetchall()
    conn.close()
    return render_template('piutang_riwayat.html', p=p, riwayat=riwayat)

@app.route('/piutang/<int:id>/hapus', methods=['POST'])
@finance_required
def piutang_hapus(id):
    conn = db()
    p = conn.execute('SELECT * FROM piutang WHERE id=?', (id,)).fetchone()
    if p:
        sisa_piutang = p['jumlah'] - p['terbayar']
        # Reverse jurnal sisa piutang yang belum terbayar: Dr 4100 (undo revenue), Cr 1120 (hapus aset)
        if sisa_piutang > 0:
            insert_jurnal(conn, str(date.today()), f"Koreksi hapus piutang: {p['pelanggan']}",
                          'KOREKSI', 'KOREKSI', [
                ('4100', sisa_piutang, 0),   # Dr Pendapatan — batalkan revenue yang belum diterima
                ('1120', 0, sisa_piutang),   # Cr Piutang — hapus saldo aset piutang
            ])
        conn.execute('DELETE FROM piutang WHERE id=?', (id,))
    conn.commit(); conn.close()
    flash('Piutang dihapus dan jurnal dikoreksi.', 'warning')
    return redirect(url_for('piutang_list'))


# ---------- HUTANG ----------
@app.route('/hutang')
@investor_required
def hutang_list():
    conn = db()
    today = date.today()
    sf = request.args.get('status','')
    q = request.args.get('q','')
    where, params = [], []
    if sf: where.append("status=?"); params.append(sf)
    if q:  where.append("(pemasok LIKE ? OR keterangan LIKE ?)"); params += [f'%{q}%',f'%{q}%']
    ws = ('WHERE '+' AND '.join(where)) if where else ''
    rows = conn.execute(f'SELECT * FROM hutang {ws} ORDER BY CASE status WHEN "LUNAS" THEN 1 ELSE 0 END, jatuh_tempo, tanggal DESC', params).fetchall()
    total = conn.execute("SELECT COALESCE(SUM(jumlah-terbayar),0) FROM hutang WHERE status!='LUNAS'").fetchone()[0]
    rows_with_status = []
    for r in rows:
        lbl, cls = piutang_status_label(r, today)
        rows_with_status.append((r, lbl, cls))
    akun_kas = conn.execute("SELECT kode,nama FROM akun WHERE is_rekening=1 ORDER BY kode").fetchall()
    conn.close()
    return render_template('hutang.html', rows=rows_with_status, total=total, sf=sf, q=q,
                           akun_kas=akun_kas, today=date.today())

@app.route('/hutang/baru', methods=['POST'])
@operator_required
def hutang_baru():
    flash('Hutang tidak bisa ditambah langsung. Catat melalui form Pengeluaran agar jurnal terbentuk.', 'warning')
    return redirect(url_for('hutang_list'))

@app.route('/hutang/<int:id>/edit', methods=['GET','POST'])
@operator_required
def hutang_edit(id):
    conn = db()
    h = conn.execute('SELECT * FROM hutang WHERE id=?', (id,)).fetchone()
    if not h:
        conn.close(); flash('Tidak ditemukan.','danger')
        return redirect(url_for('hutang_list'))
    if request.method == 'POST':
        conn.execute("""UPDATE hutang SET tanggal=?,jatuh_tempo=?,pemasok=?,keterangan=? WHERE id=?""",
            (request.form['tanggal'], request.form.get('jatuh_tempo') or None,
             request.form['pemasok'], request.form.get('keterangan',''), id))
        conn.commit(); conn.close()
        flash('Hutang diperbarui!', 'success')
        return redirect(url_for('hutang_list'))
    conn.close()
    return render_template('hutang_edit.html', h=h)

@app.route('/hutang/<int:id>/bayar', methods=['POST'])
@operator_required
def hutang_bayar(id):
    conn = db()
    jumlah = parse_rp(request.form.get('jumlah','0'))
    tanggal = request.form.get('tanggal', str(date.today()))
    catatan = request.form.get('catatan','')
    akun_kas_kode = request.form.get('akun_kas','1100')
    if not _operator_date_ok(tanggal):
        conn.close()
        flash('Operator hanya bisa mencatat pembayaran untuk bulan berjalan.', 'warning')
        return redirect(url_for('hutang_list'))
    h = conn.execute('SELECT * FROM hutang WHERE id=?', (id,)).fetchone()
    if not h:
        conn.close(); flash('Hutang tidak ditemukan.', 'danger')
        return redirect(url_for('hutang_list'))
    sisa = h['jumlah'] - h['terbayar']
    jumlah = min(jumlah, sisa)
    if jumlah <= 0:
        conn.close(); flash('Hutang sudah lunas atau jumlah tidak valid.', 'warning')
        return redirect(url_for('hutang_list'))
    conn.execute('INSERT INTO bayar_hutang(hutang_id,tanggal,jumlah,catatan) VALUES(?,?,?,?)',
                 (id, tanggal, jumlah, catatan))
    conn.execute('UPDATE hutang SET terbayar=terbayar+? WHERE id=?', (jumlah, id))
    update_hutang_status(conn, id)
    insert_jurnal(conn, tanggal, f"Pelunasan hutang: {h['pemasok']}", 'OPERASIONAL', 'PELUNASAN', [
        ('2100', jumlah, 0), (akun_kas_kode, 0, jumlah)
    ])
    conn.commit(); conn.close()
    flash('Pembayaran hutang dicatat!', 'success')
    return redirect(url_for('hutang_list'))

@app.route('/hutang/<int:id>/riwayat')
@investor_required
def hutang_riwayat(id):
    conn = db()
    h = conn.execute('SELECT * FROM hutang WHERE id=?', (id,)).fetchone()
    if not h:
        conn.close(); flash('Hutang tidak ditemukan.', 'danger')
        return redirect(url_for('hutang_list'))
    riwayat = conn.execute('SELECT * FROM bayar_hutang WHERE hutang_id=? ORDER BY tanggal DESC', (id,)).fetchall()
    conn.close()
    return render_template('hutang_riwayat.html', h=h, riwayat=riwayat)

@app.route('/hutang/<int:id>/hapus', methods=['POST'])
@finance_required
def hutang_hapus(id):
    conn = db()
    h = conn.execute('SELECT * FROM hutang WHERE id=?', (id,)).fetchone()
    if h:
        sisa_hutang = h['jumlah'] - h['terbayar']
        # Hapus sisa hutang: Dr 2100 (hapus liability), Cr 4300 (pendapatan lain — hutang dibebaskan)
        if sisa_hutang > 0:
            insert_jurnal(conn, str(date.today()), f"Koreksi hapus hutang: {h['pemasok']}",
                          'KOREKSI', 'KOREKSI', [
                ('2100', sisa_hutang, 0),  # Dr Hutang — hapus saldo liability
                ('4300', 0, sisa_hutang),  # Cr Pendapatan Lain-lain — hutang yang dibebaskan
            ])
        conn.execute('DELETE FROM hutang WHERE id=?', (id,))
    conn.commit(); conn.close()
    flash('Hutang dihapus dan jurnal dikoreksi.', 'warning')
    return redirect(url_for('hutang_list'))


# ---------- INVENTORY / STOK ----------
@app.route('/stok')
@login_required
def stok_list():
    conn = db()
    q = request.args.get('q','')
    if q:
        produk = conn.execute(
            "SELECT * FROM produk WHERE nama LIKE ? OR kode LIKE ? ORDER BY nama",
            (f'%{q}%',f'%{q}%')
        ).fetchall()
    else:
        produk = conn.execute('SELECT * FROM produk ORDER BY nama').fetchall()
    conn.close()
    return render_template('stok.html', produk=produk, q=q)

@app.route('/stok/produk/baru', methods=['POST'])
@operator_required
def produk_baru():
    conn = db()
    kode_base = request.form.get('kode','').strip()
    nama      = request.form.get('nama','').strip()
    kategori  = request.form.get('kategori','').strip()

    # multi-variant arrays
    varian_list    = request.form.getlist('varian[]')
    satuan_list    = request.form.getlist('satuan[]')
    stok_list      = request.form.getlist('stok_awal[]')
    harga_beli_list= request.form.getlist('harga_beli[]')
    harga_jual_list= request.form.getlist('harga_jual[]')
    min_stok_list  = request.form.getlist('min_stok[]')

    count = len(varian_list)
    added = 0
    errors = []
    today = date.today()

    for i in range(count):
        varian     = varian_list[i].strip() if i < len(varian_list) else ''
        satuan     = satuan_list[i] if i < len(satuan_list) else 'pcs'
        stok_awal  = float(stok_list[i] or 0) if i < len(stok_list) else 0
        harga_beli = parse_rp(harga_beli_list[i] if i < len(harga_beli_list) else '0')
        harga_jual = parse_rp(harga_jual_list[i] if i < len(harga_jual_list) else '0')
        min_stok   = float(min_stok_list[i] or 0) if i < len(min_stok_list) else 0

        # kode: base for single variant, base-1/base-2/... for multiple
        kode = kode_base if count == 1 else f'{kode_base}-{i+1}'

        try:
            conn.execute(
                'INSERT INTO produk(kode,nama,varian,satuan,harga_beli,harga_jual,stok,min_stok,kategori) '
                'VALUES(?,?,?,?,?,?,?,?,?)',
                (kode, nama, varian, satuan, harga_beli, harga_jual, stok_awal, min_stok, kategori)
            )
            if stok_awal > 0:
                pid = conn.execute('SELECT id FROM produk WHERE kode=?', (kode,)).fetchone()['id']
                conn.execute(
                    'INSERT INTO pergerakan_stok(produk_id,tanggal,jenis,qty,harga,keterangan) VALUES(?,?,?,?,?,?)',
                    (pid, today, 'MASUK', stok_awal, harga_beli, 'Stok awal')
                )
            added += 1
        except sqlite3.IntegrityError:
            errors.append(f'Kode {kode} sudah ada')

    conn.commit()
    conn.close()

    if added:
        flash(f'{added} varian produk berhasil ditambahkan!', 'success')
    for e in errors:
        flash(e, 'danger')
    return redirect(url_for('stok_list'))

@app.route('/stok/<int:id>/edit', methods=['GET','POST'])
@finance_required
def produk_edit(id):
    conn = db()
    p = conn.execute('SELECT * FROM produk WHERE id=?',(id,)).fetchone()
    if not p:
        conn.close(); flash('Produk tidak ditemukan.','danger')
        return redirect(url_for('stok_list'))
    if request.method == 'POST':
        conn.execute("""UPDATE produk SET nama=?,varian=?,satuan=?,harga_beli=?,
                        harga_jual=?,min_stok=? WHERE id=?""",
            (request.form['nama'], request.form.get('varian',''),
             request.form.get('satuan','pcs'),
             float(request.form.get('harga_beli',0) or 0),
             float(request.form.get('harga_jual',0) or 0),
             float(request.form.get('min_stok',0) or 0), id))
        conn.commit(); conn.close()
        flash('Produk diperbarui!','success')
        return redirect(url_for('stok_list'))
    conn.close()
    return render_template('produk_edit.html', p=p)

@app.route('/stok/<int:id>/gerakan', methods=['POST'])
@operator_required
def stok_gerakan(id):
    conn = db()
    jenis = request.form['jenis']
    qty   = float(request.form.get('qty',0) or 0)
    harga = float(request.form.get('harga',0) or 0)
    ket   = request.form.get('keterangan','')
    tgl   = request.form.get('tanggal', str(date.today()))
    if jenis == 'KELUAR':
        stok_kini = conn.execute('SELECT stok FROM produk WHERE id=?', (id,)).fetchone()
        if stok_kini and qty > stok_kini['stok']:
            conn.close()
            flash(f"Stok tidak mencukupi. Stok tersedia: {stok_kini['stok']:.0f}", 'danger')
            return redirect(url_for('stok_list'))
    conn.execute(
        'INSERT INTO pergerakan_stok(produk_id,tanggal,jenis,qty,harga,keterangan) VALUES(?,?,?,?,?,?)',
        (id, tgl, jenis, qty, harga, ket)
    )
    if jenis == 'MASUK':
        conn.execute('UPDATE produk SET stok=stok+? WHERE id=?', (qty,id))
    elif jenis == 'KELUAR':
        conn.execute('UPDATE produk SET stok=stok-? WHERE id=?', (qty,id))
    else:
        conn.execute('UPDATE produk SET stok=? WHERE id=?', (qty,id))
    conn.commit(); conn.close()
    flash('Pergerakan stok dicatat!','success')
    return redirect(url_for('stok_list'))

@app.route('/stok/<int:id>/riwayat')
@login_required
def stok_riwayat(id):
    conn = db()
    produk = conn.execute('SELECT * FROM produk WHERE id=?',(id,)).fetchone()
    riwayat= conn.execute(
        'SELECT * FROM pergerakan_stok WHERE produk_id=? ORDER BY tanggal DESC, id DESC',(id,)
    ).fetchall()
    conn.close()
    return render_template('stok_riwayat.html', produk=produk, riwayat=riwayat)

@app.route('/stok/<int:id>/hapus', methods=['POST'])
@finance_required
def produk_hapus(id):
    conn = db()
    conn.execute('DELETE FROM pergerakan_stok WHERE produk_id=?', (id,))
    conn.execute('DELETE FROM produk WHERE id=?', (id,))
    conn.commit(); conn.close()
    flash('Produk dihapus.','warning')
    return redirect(url_for('stok_list'))


# ---------- INVOICE ----------

def _inv_settings(conn):
    keys = ['inv_nama','inv_alamat','inv_telepon','inv_email',
            'inv_rek','inv_top','inv_top_note','inv_catatan','inv_prefix',
            'inv_logo','inv_tagline','inv_terms']
    d = {k: get_setting(conn, k, '') for k in keys}
    d['inv_nama']   = d['inv_nama']   or get_setting(conn, 'nama_usaha', 'Usaha Saya')
    d['inv_prefix'] = d['inv_prefix'] or 'INV'
    d['inv_top']    = d['inv_top']    or '30'
    try:
        d['inv_rek_list'] = json.loads(d['inv_rek']) if d['inv_rek'] else []
    except Exception:
        d['inv_rek_list'] = []
    return d

def _calc_top(inv):
    """Return actual ToP days: derived from date diff if both dates exist, else stored top_hari."""
    try:
        if inv['jatuh_tempo'] and inv['tanggal']:
            return (date.fromisoformat(inv['jatuh_tempo']) - date.fromisoformat(inv['tanggal'])).days
    except Exception:
        pass
    return inv['top_hari']

def _next_inv_number(conn, prefix):
    ym = date.today().strftime('%Y%m')
    last = conn.execute(
        "SELECT nomor FROM invoice WHERE nomor LIKE ? ORDER BY id DESC LIMIT 1",
        (f"{prefix}-{ym}-%",)
    ).fetchone()
    seq = int(last['nomor'].rsplit('-', 1)[-1]) + 1 if last else 1
    return f"{prefix}-{ym}-{seq:03d}"

@app.route('/invoice')
@investor_required
def invoice_list():
    conn = db()
    q  = request.args.get('q','')
    sf = request.args.get('status','')
    sql = "SELECT * FROM invoice WHERE 1=1"
    params = []
    if q:
        sql += " AND (pelanggan LIKE ? OR nomor LIKE ?)"
        params += [f'%{q}%', f'%{q}%']
    if sf:
        sql += " AND status=?"
        params.append(sf)
    sql += " ORDER BY tanggal DESC, id DESC"
    invoices = conn.execute(sql, params).fetchall()
    conn.close()
    return render_template('invoice_list.html', invoices=invoices, q=q, sf=sf)

@app.route('/invoice/pengaturan', methods=['GET','POST'])
@admin_required
def invoice_pengaturan():
    conn = db()
    if request.method == 'POST':
        for k in ['inv_nama','inv_alamat','inv_telepon','inv_email',
                  'inv_top','inv_top_note','inv_catatan','inv_prefix',
                  'inv_tagline','inv_terms']:
            conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)",
                         (k, request.form.get(k,'')))
        rek_list = []
        for i in range(1, 4):
            bank = request.form.get(f'rek_bank_{i}','').strip()
            norek= request.form.get(f'rek_no_{i}','').strip()
            an   = request.form.get(f'rek_an_{i}','').strip()
            if bank or norek:
                rek_list.append({'bank': bank, 'norek': norek, 'an': an})
        conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)",
                     ('inv_rek', json.dumps(rek_list)))
        # Logo upload
        logo_file = request.files.get('inv_logo_file')
        if logo_file and logo_file.filename:
            ext = logo_file.filename.rsplit('.', 1)[-1].lower()
            if ext in ALLOWED_IMG:
                for f in os.listdir(UPLOAD_FOLDER):
                    if f.startswith('inv_logo.'):
                        try: os.remove(os.path.join(UPLOAD_FOLDER, f))
                        except: pass
                logo_fn = f'inv_logo.{ext}'
                logo_file.save(os.path.join(UPLOAD_FOLDER, logo_fn))
                conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)",
                             ('inv_logo', logo_fn))
        elif request.form.get('inv_logo_delete') == '1':
            old = get_setting(conn, 'inv_logo', '')
            if old:
                try: os.remove(os.path.join(UPLOAD_FOLDER, old))
                except: pass
            conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)",
                         ('inv_logo', ''))
        conn.commit(); conn.close()
        flash('Template invoice disimpan!', 'success')
        return redirect(url_for('invoice_pengaturan'))
    inv_s = _inv_settings(conn)
    conn.close()
    return render_template('invoice_pengaturan.html', inv_s=inv_s)

def _save_invoice_items(conn, inv_id, form):
    conn.execute("DELETE FROM invoice_item WHERE invoice_id=?", (inv_id,))
    desks = form.getlist('deskripsi[]')
    qtys  = form.getlist('qty[]')
    sats  = form.getlist('satuan[]')
    hargs = form.getlist('harga_satuan[]')
    diss  = form.getlist('diskon_item[]')
    for i, desk in enumerate(desks):
        if not desk.strip(): continue
        qty  = float(qtys[i]  if i < len(qtys)  else 1) or 1
        sat  = sats[i]        if i < len(sats)  else 'pcs'
        harg = float(str(hargs[i]).replace('.','').replace(',','.') if i < len(hargs) else 0) or 0
        dis_i= float(str(diss[i]).replace('.','').replace(',','.') if i < len(diss)  else 0) or 0
        conn.execute(
            "INSERT INTO invoice_item(invoice_id,deskripsi,qty,satuan,harga_satuan,diskon_item,subtotal) VALUES(?,?,?,?,?,?,?)",
            (inv_id, desk.strip(), qty, sat, harg, dis_i, qty * harg - dis_i))

@app.route('/invoice/preview')
@investor_required
def invoice_preview():
    conn = db()
    inv_s = _inv_settings(conn)
    conn.close()
    today_str = date.today().strftime('%Y-%m-%d')
    top = int(inv_s['inv_top'] or 30)
    jt  = (date.today() + timedelta(days=top)).strftime('%Y-%m-%d')
    inv = {
        'nomor': f"{inv_s['inv_prefix']}-{date.today().strftime('%Y%m')}-001",
        'tanggal': today_str,
        'jatuh_tempo': jt,
        'top_hari': top,
        'top_note': inv_s['inv_top_note'],
        'pelanggan': 'Contoh Pelanggan / PT Maju Bersama',
        'alamat_pelanggan': 'Jl. Contoh No. 123, Jakarta',
        'telepon_pelanggan': '0812-3456-7890',
        'diskon': 50000,
        'ongkir': 25000,
        'biaya_lain': 0,
        'catatan': '',
        'status': 'DRAFT',
    }
    items = [
        {'deskripsi': 'Produk / Jasa Contoh A', 'qty': 2, 'satuan': 'pcs',
         'harga_satuan': 500000, 'diskon_item': 0, 'subtotal': 1000000},
        {'deskripsi': 'Produk / Jasa Contoh B', 'qty': 1, 'satuan': 'unit',
         'harga_satuan': 750000, 'diskon_item': 25000, 'subtotal': 725000},
    ]
    subtotal = sum(it['subtotal'] for it in items)
    total    = max(0, subtotal - inv['diskon'] + inv['ongkir'] + inv['biaya_lain'])
    return render_template('invoice_print.html', inv=inv, items=items, inv_s=inv_s,
                           subtotal=subtotal, total=total, autoprint=False, preview=True,
                           actual_top=top)

@app.route('/invoice/baru', methods=['GET','POST'])
@operator_required
def invoice_baru():
    conn = db()
    inv_s = _inv_settings(conn)
    produk_list = conn.execute("SELECT * FROM produk ORDER BY nama").fetchall()
    if request.method == 'POST':
        prefix = inv_s['inv_prefix']
        nomor  = request.form.get('nomor','').strip() or _next_inv_number(conn, prefix)
        if conn.execute("SELECT id FROM invoice WHERE nomor=?", (nomor,)).fetchone():
            flash(f'Nomor invoice {nomor} sudah digunakan.', 'danger')
            conn.close()
            return redirect(url_for('invoice_baru'))
        top_hari = int(request.form.get('top_hari') or inv_s['inv_top'] or 30)
        top_note = request.form.get('top_note', inv_s['inv_top_note']).strip()
        tanggal  = request.form['tanggal']
        jt = request.form.get('jatuh_tempo','').strip() or \
             (date.fromisoformat(tanggal) + timedelta(days=top_hari)).strftime('%Y-%m-%d')
        inv_id = conn.execute(
            "INSERT INTO invoice(nomor,tanggal,jatuh_tempo,top_hari,top_note,pelanggan,alamat_pelanggan,telepon_pelanggan,diskon,ongkir,biaya_lain,catatan) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
            (nomor, tanggal, jt, top_hari, top_note,
             request.form.get('pelanggan',''),
             request.form.get('alamat_pelanggan',''),
             request.form.get('telepon_pelanggan',''),
             parse_rp(request.form.get('diskon','0')),
             parse_rp(request.form.get('ongkir','0')),
             parse_rp(request.form.get('biaya_lain','0')),
             request.form.get('catatan',''))
        ).lastrowid
        _save_invoice_items(conn, inv_id, request.form)
        add_log(conn, 'Invoice dibuat', f"{nomor} | {request.form.get('pelanggan','')}", 'INPUT')
        conn.commit(); conn.close()
        flash(f'Invoice {nomor} berhasil dibuat!', 'success')
        return redirect(url_for('invoice_view', id=inv_id))
    next_num = _next_inv_number(conn, inv_s['inv_prefix'])
    conn.close()
    return render_template('invoice_form.html', inv_s=inv_s, produk_list=produk_list,
                           next_num=next_num, today=date.today().strftime('%Y-%m-%d'),
                           edit=None, items=[])

@app.route('/invoice/<int:id>/edit', methods=['GET','POST'])
@operator_required
def invoice_edit(id):
    conn = db()
    inv = conn.execute("SELECT * FROM invoice WHERE id=?", (id,)).fetchone()
    if not inv:
        conn.close(); flash('Invoice tidak ditemukan.','danger')
        return redirect(url_for('invoice_list'))
    inv_s = _inv_settings(conn)
    produk_list = conn.execute("SELECT * FROM produk ORDER BY nama").fetchall()
    if request.method == 'POST':
        if inv['status'] == 'PAID' and session.get('role') not in ('ADMIN', 'FINANCE'):
            conn.close(); flash('Invoice yang sudah lunas hanya bisa diedit oleh Finance/Admin.', 'danger')
            return redirect(url_for('invoice_view', id=id))
        nomor_e = request.form.get('nomor', inv['nomor']).strip() or inv['nomor']
        if nomor_e != inv['nomor'] and conn.execute("SELECT id FROM invoice WHERE nomor=? AND id!=?", (nomor_e, id)).fetchone():
            conn.close(); flash(f'Nomor invoice {nomor_e} sudah digunakan.', 'danger')
            return redirect(url_for('invoice_edit', id=id))
        top_hari_e = int(request.form.get('top_hari') or inv_s['inv_top'] or 30)
        top_note_e = request.form.get('top_note', '').strip()
        tanggal_e  = request.form['tanggal']
        jt_e = request.form.get('jatuh_tempo','').strip() or \
               (date.fromisoformat(tanggal_e) + timedelta(days=top_hari_e)).strftime('%Y-%m-%d')
        conn.execute("""UPDATE invoice SET nomor=?,tanggal=?,jatuh_tempo=?,top_hari=?,top_note=?,
                        pelanggan=?,alamat_pelanggan=?,telepon_pelanggan=?,
                        diskon=?,ongkir=?,biaya_lain=?,catatan=? WHERE id=?""",
                     (nomor_e, tanggal_e, jt_e, top_hari_e, top_note_e,
                      request.form.get('pelanggan',''),
                      request.form.get('alamat_pelanggan',''),
                      request.form.get('telepon_pelanggan',''),
                      parse_rp(request.form.get('diskon','0')),
                      parse_rp(request.form.get('ongkir','0')),
                      parse_rp(request.form.get('biaya_lain','0')),
                      request.form.get('catatan',''), id))
        _save_invoice_items(conn, id, request.form)
        conn.commit(); conn.close()
        flash('Invoice diperbarui!', 'success')
        return redirect(url_for('invoice_view', id=id))
    items = conn.execute("SELECT * FROM invoice_item WHERE invoice_id=? ORDER BY id", (id,)).fetchall()
    conn.close()
    return render_template('invoice_form.html', inv_s=inv_s, produk_list=produk_list,
                           today=date.today().strftime('%Y-%m-%d'), edit=inv, items=items)

@app.route('/invoice/<int:id>')
@investor_required
def invoice_view(id):
    conn = db()
    inv   = conn.execute("SELECT * FROM invoice WHERE id=?", (id,)).fetchone()
    if not inv:
        conn.close(); flash('Invoice tidak ditemukan.','danger')
        return redirect(url_for('invoice_list'))
    items = conn.execute("SELECT * FROM invoice_item WHERE invoice_id=? ORDER BY id", (id,)).fetchall()
    inv_s = _inv_settings(conn)
    conn.close()
    subtotal  = sum(it['subtotal'] for it in items)
    total     = max(0, subtotal - inv['diskon'] + inv['ongkir'] + inv['biaya_lain'])
    actual_top = _calc_top(inv)
    return render_template('invoice_view.html', inv=inv, items=items, inv_s=inv_s,
                           subtotal=subtotal, total=total, actual_top=actual_top)

@app.route('/invoice/<int:id>/cetak')
@investor_required
def invoice_cetak(id):
    conn = db()
    inv   = conn.execute("SELECT * FROM invoice WHERE id=?", (id,)).fetchone()
    if not inv:
        conn.close(); return redirect(url_for('invoice_list'))
    items = conn.execute("SELECT * FROM invoice_item WHERE invoice_id=? ORDER BY id", (id,)).fetchall()
    inv_s = _inv_settings(conn)
    subtotal = sum(it['subtotal'] for it in items)
    total    = max(0, subtotal - inv['diskon'] + inv['ongkir'] + inv['biaya_lain'])

    # Cari piutang terkait invoice ini
    # Prioritas 1: keterangan mengandung nomor invoice (untuk invoice baru)
    piutang_row = conn.execute(
        "SELECT * FROM piutang WHERE pelanggan=? AND keterangan LIKE ? ORDER BY id DESC LIMIT 1",
        (inv['pelanggan'], f"%{inv['nomor']}%")
    ).fetchone()
    # Prioritas 2: tanggal + pelanggan sama, jumlah <= total invoice (piutang partial/full)
    if not piutang_row:
        piutang_row = conn.execute(
            """SELECT * FROM piutang WHERE pelanggan=? AND tanggal=?
               AND jumlah <= ? AND jumlah > 0
               ORDER BY ABS(jumlah - ?) LIMIT 1""",
            (inv['pelanggan'], inv['tanggal'], total + 1, total)
        ).fetchone()
    # Prioritas 3: pelanggan + jumlah tepat sama (tanpa batasan tanggal)
    if not piutang_row:
        piutang_row = conn.execute(
            "SELECT * FROM piutang WHERE pelanggan=? AND ROUND(jumlah,0)=ROUND(?,0) ORDER BY id DESC LIMIT 1",
            (inv['pelanggan'], total)
        ).fetchone()

    # Nilai STATIS dari transaksi awal — tidak berubah meski ada pembayaran berikutnya
    if piutang_row:
        p_jumlah     = float(piutang_row['jumlah'] or 0)
        pay_dp       = float(total) - p_jumlah   # uang muka dibayar saat transaksi
        pay_sisa     = p_jumlah                  # sisa tagihan tetap (statis)
        # Stamp dinamis — boleh berubah sesuai status piutang saat ini
        pay_stamp    = piutang_row['status']     # 'LUNAS' / 'SEBAGIAN' / 'BELUM LUNAS'
        show_payment = True
    elif inv['status'] == 'PAID':
        pay_dp       = float(total)
        pay_sisa     = 0.0
        pay_stamp    = 'LUNAS'
        show_payment = True
    elif inv['jatuh_tempo']:
        pay_dp       = 0.0
        pay_sisa     = float(total)
        pay_stamp    = 'BELUM LUNAS'
        show_payment = True
    else:
        pay_dp = pay_sisa = 0.0
        pay_stamp    = None
        show_payment = False

    conn.close()
    autoprint  = request.args.get('autoprint','0') == '1'
    actual_top = _calc_top(inv)
    return render_template('invoice_print.html', inv=inv, items=items, inv_s=inv_s,
                           subtotal=subtotal, total=total, autoprint=autoprint,
                           actual_top=actual_top,
                           show_payment=show_payment,
                           pay_dp=pay_dp, pay_sisa=pay_sisa, pay_stamp=pay_stamp)

@app.route('/invoice/<int:id>/status', methods=['POST'])
@operator_required
def invoice_status_update(id):
    new_status = request.form.get('status', 'SENT')
    # Hanya ADMIN dan FINANCE yang boleh set status PAID
    if new_status == 'PAID' and session.get('role') not in ('ADMIN', 'FINANCE'):
        flash('Hanya Admin/Finance yang bisa menandai invoice sebagai Lunas.', 'danger')
        return redirect(url_for('invoice_view', id=id))
    conn = db()
    conn.execute("UPDATE invoice SET status=? WHERE id=?", (new_status, id))
    conn.commit(); conn.close()
    return redirect(url_for('invoice_view', id=id))

@app.route('/invoice/<int:id>/hapus', methods=['POST'])
@finance_required
def invoice_hapus(id):
    conn = db()
    conn.execute("DELETE FROM invoice WHERE id=?", (id,))
    conn.commit(); conn.close()
    flash('Invoice dihapus.', 'info')
    return redirect(url_for('invoice_list'))


# ---------- LAPORAN ----------
@app.route('/neraca')
@investor_required
def neraca():
    conn = db()
    ed = request.args.get('ed', date.today().strftime('%Y-%m-%d'))
    rows = conn.execute("""
        SELECT a.kode, a.nama, a.tipe, a.subtipe, a.saldo_normal,
               COALESCE(SUM(d.debit),0) as td, COALESCE(SUM(d.kredit),0) as tk
        FROM akun a
        LEFT JOIN (
            SELECT dj.akun_id, dj.debit, dj.kredit
            FROM detail_jurnal dj
            JOIN jurnal j ON j.id = dj.jurnal_id AND j.tanggal <= ?
        ) d ON d.akun_id = a.id
        GROUP BY a.id ORDER BY a.kode
    """, (ed,)).fetchall()
    aset={};liab={};ekuitas_list=[]
    total_aset=total_liab=total_ekuitas=0
    for r in rows:
        saldo = (r['td']-r['tk']) if r['saldo_normal']=='DEBIT' else (r['tk']-r['td'])
        item = dict(kode=r['kode'],nama=r['nama'],saldo=saldo)
        if r['tipe']=='ASET':
            # Contra-asset (saldo normal KREDIT, mis. Akumulasi Penyusutan) mengurangi aset
            aset_saldo = -saldo if r['saldo_normal'] == 'KREDIT' else saldo
            aset.setdefault(r['subtipe'] or 'Aset Lancar',[]).append(
                dict(kode=r['kode'], nama=r['nama'], saldo=aset_saldo))
            total_aset += aset_saldo
        elif r['tipe']=='LIABILITAS':
            liab.setdefault(r['subtipe'] or 'Liabilitas Lancar',[]).append(item)
            total_liab += saldo
        elif r['tipe']=='EKUITAS':
            # Akun ekuitas berdebet (contoh: 3300 Prive) MENGURANGI ekuitas
            ekuitas_saldo = -saldo if r['saldo_normal'] == 'DEBIT' else saldo
            ekuitas_list.append(dict(kode=r['kode'], nama=r['nama'], saldo=ekuitas_saldo))
            total_ekuitas += ekuitas_saldo
    fiscal_start  = f"{ed[:4]}-01-01"
    laba_tahun    = calc_profitability(conn, fiscal_start, ed)['laba_bersih']
    laba_all      = calc_profitability(conn, '2000-01-01', ed)['laba_bersih']
    laba_ditahan  = laba_all - laba_tahun
    if abs(laba_ditahan) > 0.01:
        ekuitas_list.append(dict(kode='-', nama='Laba Ditahan (Akumulasi)', saldo=laba_ditahan))
        total_ekuitas += laba_ditahan
    ekuitas_list.append(dict(kode='-', nama=f'Laba Bersih Tahun {ed[:4]}', saldo=laba_tahun))
    total_ekuitas += laba_tahun

    # Diagnostik 1: jurnal yang tidak seimbang (debit ≠ kredit)
    unbalanced = conn.execute("""
        SELECT j.id, j.tanggal, j.keterangan, j.tipe_tx,
               ROUND(SUM(dj.debit),2)  AS total_debit,
               ROUND(SUM(dj.kredit),2) AS total_kredit,
               ROUND(SUM(dj.debit) - SUM(dj.kredit), 2) AS selisih
        FROM jurnal j
        JOIN detail_jurnal dj ON dj.jurnal_id = j.id
        GROUP BY j.id
        HAVING ABS(SUM(dj.debit) - SUM(dj.kredit)) > 0.01
        ORDER BY ABS(SUM(dj.debit) - SUM(dj.kredit)) DESC
        LIMIT 20
    """).fetchall()
    # Diagnostik 2: entri jurnal yang merujuk akun yang sudah dihapus (yatim)
    orphaned = conn.execute("""
        SELECT j.id, j.tanggal, j.keterangan, j.tipe_tx,
               ROUND(SUM(dj.debit),2)  AS orphan_debit,
               ROUND(SUM(dj.kredit),2) AS orphan_kredit,
               GROUP_CONCAT(dj.akun_id) AS akun_ids_hilang
        FROM detail_jurnal dj
        LEFT JOIN akun a ON a.id = dj.akun_id
        JOIN jurnal j ON j.id = dj.jurnal_id
        WHERE a.id IS NULL
        GROUP BY j.id
        ORDER BY j.tanggal DESC
        LIMIT 20
    """).fetchall()
    # Cek total global
    global_check = conn.execute(
        "SELECT ROUND(SUM(debit)-SUM(kredit),2) as diff FROM detail_jurnal"
    ).fetchone()
    global_diff = float(global_check['diff'] or 0)

    conn.close()
    return render_template('neraca.html',
        aset=aset,liab=liab,ekuitas_list=ekuitas_list,
        total_aset=total_aset,total_liab=total_liab,total_ekuitas=total_ekuitas,ed=ed,
        unbalanced=unbalanced, orphaned=orphaned, global_diff=global_diff)

@app.route('/laba-rugi')
@investor_required
def laba_rugi():
    conn = db()
    today = date.today()
    sd = request.args.get('sd', today.replace(day=1).strftime('%Y-%m-%d'))
    ed = request.args.get('ed', today.strftime('%Y-%m-%d'))
    pnl = calc_profitability(conn, sd, ed)
    rows = conn.execute("""
        SELECT a.kode, a.nama, a.tipe, a.subtipe,
               COALESCE(SUM(CASE WHEN a.saldo_normal='DEBIT' THEN d.debit-d.kredit
                                 ELSE d.kredit-d.debit END),0) as saldo
        FROM akun a
        LEFT JOIN (
            SELECT dj.akun_id, dj.debit, dj.kredit
            FROM detail_jurnal dj
            JOIN jurnal j ON j.id = dj.jurnal_id AND j.tanggal >= ? AND j.tanggal <= ?
        ) d ON d.akun_id = a.id
        WHERE a.tipe IN ('PENDAPATAN','BEBAN')
        GROUP BY a.id ORDER BY a.kode
    """, (sd, ed)).fetchall()
    pendapatan={}; beban={}
    for r in rows:
        item = dict(kode=r['kode'],nama=r['nama'],saldo=r['saldo'])
        if r['tipe']=='PENDAPATAN':
            pendapatan.setdefault(r['subtipe'] or 'Pendapatan',[]).append(item)
        else:
            beban.setdefault(r['subtipe'] or 'Beban',[]).append(item)
    conn.close()
    return render_template('laba_rugi.html',
        pendapatan=pendapatan, beban=beban, pnl=pnl, sd=sd, ed=ed)

@app.route('/buku-besar')
@investor_required
def buku_besar():
    conn  = db()
    today = date.today()
    sd    = request.args.get('sd', today.replace(day=1).strftime('%Y-%m-%d'))
    ed    = request.args.get('ed', today.strftime('%Y-%m-%d'))
    kode  = request.args.get('akun', '')

    # Semua akun untuk dropdown (grouped by tipe)
    semua_akun = conn.execute(
        "SELECT kode, nama, tipe, saldo_normal FROM akun ORDER BY kode"
    ).fetchall()

    entri = []
    akun_info   = None
    saldo_awal  = 0
    total_debit = 0
    total_kredit = 0

    if kode:
        akun_row = conn.execute(
            "SELECT id, kode, nama, tipe, saldo_normal FROM akun WHERE kode=?", (kode,)
        ).fetchone()

        if akun_row:
            akun_info = dict(akun_row)
            aid = akun_row['id']
            sn  = akun_row['saldo_normal']   # 'DEBIT' atau 'KREDIT'

            # ── Saldo awal: semua entry SEBELUM periode ──────────────────
            r = conn.execute("""
                SELECT COALESCE(SUM(d.debit),0) as tot_d,
                       COALESCE(SUM(d.kredit),0) as tot_k
                FROM detail_jurnal d
                JOIN jurnal j ON j.id = d.jurnal_id
                WHERE d.akun_id = ? AND j.tanggal < ?
            """, [aid, sd]).fetchone()
            tot_d_awal  = r['tot_d']
            tot_k_awal  = r['tot_k']
            if sn == 'DEBIT':
                saldo_awal = tot_d_awal - tot_k_awal
            else:
                saldo_awal = tot_k_awal - tot_d_awal

            # ── Entri dalam periode ──────────────────────────────────────
            rows = conn.execute("""
                SELECT j.id, j.tanggal, j.keterangan, j.referensi, j.tipe_tx,
                       d.debit, d.kredit
                FROM detail_jurnal d
                JOIN jurnal j ON j.id = d.jurnal_id
                WHERE d.akun_id = ? AND j.tanggal >= ? AND j.tanggal <= ?
                ORDER BY j.tanggal ASC, j.id ASC
            """, [aid, sd, ed]).fetchall()

            saldo_berjalan = saldo_awal
            for row in rows:
                dbt = row['debit']
                krd = row['kredit']
                if sn == 'DEBIT':
                    saldo_berjalan += dbt - krd
                else:
                    saldo_berjalan += krd - dbt
                total_debit  += dbt
                total_kredit += krd
                entri.append({
                    'jurnal_id'  : row['id'],
                    'tanggal'    : row['tanggal'],
                    'keterangan' : row['keterangan'],
                    'referensi'  : row['referensi'] or '',
                    'tipe_tx'    : row['tipe_tx'],
                    'debit'      : dbt,
                    'kredit'     : krd,
                    'saldo'      : saldo_berjalan,
                })

    saldo_akhir = saldo_awal + (total_debit - total_kredit) if (akun_info and akun_info['saldo_normal']=='DEBIT') \
                  else saldo_awal + (total_kredit - total_debit)

    conn.close()
    return render_template('buku_besar.html',
        semua_akun=semua_akun, akun_info=akun_info,
        entri=entri, sd=sd, ed=ed, kode=kode,
        saldo_awal=saldo_awal, saldo_akhir=saldo_akhir,
        total_debit=total_debit, total_kredit=total_kredit,
        today=today.strftime('%Y-%m-%d'),
    )


@app.route('/arus-kas')
@investor_required
def arus_kas():
    conn = db()
    today = date.today()
    sd = request.args.get('sd', today.replace(day=1).strftime('%Y-%m-%d'))
    ed = request.args.get('ed', today.strftime('%Y-%m-%d'))
    kas_ids = get_rekening_ids(conn)
    saldo_awal = 0; operasional=[]; investasi=[]; pendanaan=[]; total_op=total_inv=total_pend=0
    if kas_ids:
        ph = ','.join('?'*len(kas_ids))
        saldo_awal = conn.execute(f"""
            SELECT COALESCE(SUM(d.debit-d.kredit),0)
            FROM detail_jurnal d JOIN jurnal j ON j.id=d.jurnal_id
            WHERE d.akun_id IN ({ph}) AND j.tanggal<?
        """, kas_ids+[sd]).fetchone()[0]
        flows = conn.execute(f"""
            SELECT j.id,j.tanggal,j.keterangan,j.kategori,
                   SUM(CASE WHEN d.akun_id IN ({ph}) THEN d.debit-d.kredit ELSE 0 END) as net_kas
            FROM jurnal j JOIN detail_jurnal d ON d.jurnal_id=j.id
            WHERE j.tanggal>=? AND j.tanggal<=?
            GROUP BY j.id HAVING net_kas!=0 ORDER BY j.tanggal,j.id
        """, kas_ids+[sd,ed]).fetchall()
        operasional = [r for r in flows if r['kategori']=='OPERASIONAL']
        investasi   = [r for r in flows if r['kategori']=='INVESTASI']
        pendanaan   = [r for r in flows if r['kategori']=='PENDANAAN']
        total_op   = sum(r['net_kas'] for r in operasional)
        total_inv  = sum(r['net_kas'] for r in investasi)
        total_pend = sum(r['net_kas'] for r in pendanaan)
    saldo_akhir = saldo_awal + total_op + total_inv + total_pend
    cf = calc_cashflow(conn, sd, ed)
    conn.close()
    return render_template('arus_kas.html',
        operasional=operasional,investasi=investasi,pendanaan=pendanaan,
        saldo_awal=saldo_awal,total_op=total_op,total_inv=total_inv,
        total_pend_fin=total_pend,saldo_akhir=saldo_akhir,sd=sd,ed=ed,
        uang_masuk=cf['masuk'],uang_keluar=cf['keluar'])


# ---------- CHARTS ----------
@app.route('/charts')
@investor_required
def charts():
    months = int(request.args.get('months', 6))
    return render_template('charts.html', months=months)

@app.route('/api/chart-data')
@investor_required
def api_chart_data():
    n = int(request.args.get('months', 12))
    today = date.today()
    conn = db()
    data = []
    kas_ids = get_rekening_ids(conn)
    kas_ph  = ','.join('?'*len(kas_ids)) if kas_ids else '0'

    for i in range(n-1, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0: m += 12; y -= 1
        ms = date(y, m, 1)
        me = month_end(y, m)
        sd = ms.strftime('%Y-%m-%d')
        ed = me.strftime('%Y-%m-%d')
        label = f"{BULAN[m-1]} {y}"

        pnl = calc_profitability(conn, sd, ed)
        cf  = calc_cashflow(conn, sd, ed)

        saldo_kum = conn.execute(f"""
            SELECT COALESCE(SUM(d.debit-d.kredit),0)
            FROM detail_jurnal d JOIN jurnal j ON j.id=d.jurnal_id
            WHERE d.akun_id IN ({kas_ph}) AND j.tanggal<=?
        """, kas_ids+[ed]).fetchone()[0] if kas_ids else 0

        data.append({
            'label': label,
            'pendapatan': pnl['rev'], 'hpp': pnl['hpp'],
            'laba_kotor': pnl['laba_kotor'], 'op_exp': pnl['op_exp'],
            'laba_op': pnl['laba_op'], 'depr': pnl['depr'],
            'ebit': pnl['ebit'], 'tax': pnl['tax'],
            'laba_bersih': pnl['laba_bersih'], 'prive': pnl['prive'],
            'laba_tahan': pnl['laba_tahan'],
            'uang_masuk': cf['masuk'], 'uang_keluar': cf['keluar'],
            'saldo_kumulatif': saldo_kum,
        })
    conn.close()
    return jsonify(data)


@app.route('/api/chart-year')
@investor_required
def api_chart_year():
    """12 bulan untuk tahun yang dipilih — data lengkap untuk 5 metrik chart."""
    today = date.today()
    y     = int(request.args.get('year', today.year))
    conn  = db()

    kas_ids = get_rekening_ids(conn)
    kas_ph  = ','.join('?'*len(kas_ids)) if kas_ids else '0'

    data = []
    for m in range(1, 13):
        ms  = date(y, m, 1).strftime('%Y-%m-%d')
        me  = month_end(y, m).strftime('%Y-%m-%d')
        pnl = calc_profitability(conn, ms, me)
        total_beban = pnl['hpp'] + pnl['op_exp'] + pnl['depr'] + pnl['interest'] + pnl['tax']

        # Arus kas: debit/kredit pada akun kas per bulan
        if kas_ids:
            kas_masuk  = conn.execute(
                f"SELECT COALESCE(SUM(d.debit-d.kredit),0) FROM detail_jurnal d "
                f"JOIN jurnal j ON j.id=d.jurnal_id "
                f"WHERE d.akun_id IN ({kas_ph}) AND j.tanggal>=? AND j.tanggal<=? AND j.tipe_tx='PEMASUKAN'",
                kas_ids+[ms, me]).fetchone()[0] or 0
            kas_keluar = conn.execute(
                f"SELECT COALESCE(SUM(d.kredit-d.debit),0) FROM detail_jurnal d "
                f"JOIN jurnal j ON j.id=d.jurnal_id "
                f"WHERE d.akun_id IN ({kas_ph}) AND j.tanggal>=? AND j.tanggal<=? AND j.tipe_tx='PENGELUARAN'",
                kas_ids+[ms, me]).fetchone()[0] or 0
        else:
            kas_masuk = kas_keluar = 0

        rev = pnl['rev'] or 0
        gm_pct = round(pnl['laba_kotor'] / rev * 100, 1) if rev else 0
        nm_pct = round(pnl['laba_bersih'] / rev * 100, 1) if rev else 0

        data.append({
            'label':       BULAN[m-1],
            'month':       m,
            'is_future':   (y == today.year and m > today.month) or y > today.year,
            # Chart 1 – Pendapatan vs Beban
            'pendapatan':  rev,
            'total_beban': total_beban,
            'laba_bersih': pnl['laba_bersih'],
            'laba_tahan':  pnl['laba_tahan'],
            # Chart 2 – Laba Bersih (uses laba_bersih)
            # Chart 3 – Arus Kas
            'kas_masuk':   kas_masuk,
            'kas_keluar':  kas_keluar,
            'net_kas':     kas_masuk - kas_keluar,
            # Chart 4 – Margin %
            'gm_pct':      gm_pct,
            'nm_pct':      nm_pct,
            # Chart 5 – Breakdown Biaya
            'hpp':         pnl['hpp'],
            'op_exp':      pnl['op_exp'],
            'depr':        pnl['depr'],
            'tax':         pnl['tax'],
        })
    conn.close()
    return jsonify(data)

@app.route('/api/chart-tahunan')
@investor_required
def api_chart_tahunan():
    n = int(request.args.get('years', 5))
    today = date.today()
    conn = db()
    kas_ids = get_rekening_ids(conn)
    kas_ph  = ','.join('?'*len(kas_ids)) if kas_ids else '0'
    data = []

    for i in range(n-1, -1, -1):
        y = today.year - i
        sd = f"{y}-01-01"
        ed = f"{y}-12-31"

        pnl = calc_profitability(conn, sd, ed)
        cf  = calc_cashflow(conn, sd, ed)

        saldo_akhir = conn.execute(f"""
            SELECT COALESCE(SUM(d.debit-d.kredit),0)
            FROM detail_jurnal d JOIN jurnal j ON j.id=d.jurnal_id
            WHERE d.akun_id IN ({kas_ph}) AND j.tanggal<=?
        """, kas_ids+[ed]).fetchone()[0] if kas_ids else 0

        total_beban = pnl['hpp'] + pnl['op_exp'] + pnl['depr'] + pnl['tax'] + pnl['other_e']

        data.append({
            'label': str(y),
            'pendapatan': pnl['rev'], 'hpp': pnl['hpp'],
            'total_beban': total_beban, 'op_exp': pnl['op_exp'],
            'laba_kotor': pnl['laba_kotor'], 'laba_op': pnl['laba_op'],
            'laba_bersih': pnl['laba_bersih'], 'laba_tahan': pnl['laba_tahan'],
            'depr': pnl['depr'], 'tax': pnl['tax'], 'prive': pnl['prive'],
            'uang_masuk': cf['masuk'], 'uang_keluar': cf['keluar'],
            'saldo_akhir': saldo_akhir,
            'pct_laba_bersih': pnl['pct_laba_bersih'],
        })

    # Hitung pertumbuhan pendapatan YoY (%)
    for i in range(len(data)):
        prev = data[i-1]['pendapatan'] if i > 0 else 0
        curr = data[i]['pendapatan']
        data[i]['growth_pct'] = round((curr - prev) / prev * 100, 1) if prev else 0

    conn.close()
    return jsonify(data)


# ---------- DAFTAR TRANSAKSI (Semua / Pemasukan / Pengeluaran) ----------
@app.route('/daftar-transaksi')
@login_required
def daftar_transaksi():
    conn = db()
    tab       = request.args.get('tab', 'semua')
    beban     = request.args.get('beban', '')
    q         = request.args.get('q', '')
    _today    = date.today()
    _def_from = session.get('dash_sd', _today.replace(day=1).strftime('%Y-%m-%d'))
    _def_to   = session.get('dash_ed', _today.strftime('%Y-%m-%d'))
    date_from = request.args.get('date_from', _def_from)
    date_to   = request.args.get('date_to',   _def_to)

    q_where  = ["j.keterangan LIKE ?"] if q else []
    q_params = [f'%{q}%'] if q else []
    if date_from:
        q_where.append("j.tanggal >= ?"); q_params.append(date_from)
    if date_to:
        q_where.append("j.tanggal <= ?"); q_params.append(date_to)

    # Sub-label SQL: tampilkan kategori bahasa UMKM bukan akuntansi
    _sub_label_sql = """
        CASE
          WHEN j.tipe_tx='PEMASUKAN' THEN 'Penjualan'
          WHEN j.kategori='INVESTASI' THEN 'Investasi Aset'
          WHEN j.kategori='PENDANAAN' THEN 'Penarikan Owner'
          WHEN EXISTS(SELECT 1 FROM detail_jurnal dj JOIN akun ax ON ax.id=dj.akun_id
                      WHERE dj.jurnal_id=j.id AND ax.kode='1130' AND dj.debit>0)
               THEN 'Bahan Baku'
          WHEN EXISTS(SELECT 1 FROM detail_jurnal dj JOIN akun ax ON ax.id=dj.akun_id
                      WHERE dj.jurnal_id=j.id AND ax.kode='6170' AND dj.debit>0)
               THEN 'Pajak'
          ELSE 'Operasional'
        END as sub_label
    """

    def _fetch_masuk(extra_where=[], extra_params=[]):
        w = ["j.tipe_tx='PEMASUKAN'"] + extra_where
        ws = 'WHERE ' + ' AND '.join(w)
        return conn.execute(f"""
            SELECT j.id, j.tanggal, j.keterangan, j.kategori, 'PEMASUKAN' as tipe_tx,
                   (SELECT COALESCE(SUM(dj2.kredit),0)
                    FROM detail_jurnal dj2 JOIN akun a2 ON a2.id=dj2.akun_id
                    WHERE dj2.jurnal_id=j.id AND a2.kode LIKE '4%') as total,
                   {_sub_label_sql}
            FROM jurnal j {ws} ORDER BY j.tanggal DESC, j.id DESC
        """, extra_params).fetchall()

    def _fetch_keluar(extra_where=[], extra_params=[]):
        w = ["j.tipe_tx='PENGELUARAN'"] + extra_where
        ws = 'WHERE ' + ' AND '.join(w)
        return conn.execute(f"""
            SELECT j.id, j.tanggal, j.keterangan, j.kategori, 'PENGELUARAN' as tipe_tx,
                   COALESCE(SUM(d.debit),0) as total,
                   {_sub_label_sql}
            FROM jurnal j LEFT JOIN detail_jurnal d ON d.jurnal_id=j.id
            {ws} GROUP BY j.id ORDER BY j.tanggal DESC, j.id DESC
        """, extra_params).fetchall()

    if tab == 'semua':
        rp = [dict(r) for r in _fetch_masuk(q_where, q_params)]
        rk = [dict(r) for r in _fetch_keluar(q_where, q_params)]
        rows = sorted(rp + rk, key=lambda r: (r['tanggal'], r['id']), reverse=True)
        total_masuk  = sum(r['total'] for r in rp)
        total_keluar = sum(r['total'] for r in rk)
        total_all    = total_masuk - total_keluar
    elif tab == 'pemasukan':
        rows = _fetch_masuk(q_where, q_params)
        total_masuk = total_all = sum(r['total'] for r in rows)
        total_keluar = 0
    else:
        beban_w = []
        if beban == 'bahan_baku':
            # Bahan baku → Dr 1130 (Persediaan), bukan 5100
            beban_w.append("EXISTS(SELECT 1 FROM detail_jurnal dj JOIN akun a ON a.id=dj.akun_id WHERE dj.jurnal_id=j.id AND a.kode='1130' AND dj.debit>0)")
        elif beban == 'operasional':
            beban_w += ["j.kategori='OPERASIONAL'",
                        "NOT EXISTS(SELECT 1 FROM detail_jurnal dj JOIN akun a ON a.id=dj.akun_id WHERE dj.jurnal_id=j.id AND a.kode IN('1130','6170') AND dj.debit>0)"]
        elif beban == 'investasi':
            beban_w.append("j.kategori='INVESTASI'")
        elif beban == 'pajak':
            beban_w.append("EXISTS(SELECT 1 FROM detail_jurnal dj JOIN akun a ON a.id=dj.akun_id WHERE dj.jurnal_id=j.id AND a.kode='6170' AND dj.debit>0)")
        elif beban == 'penarikan':
            beban_w.append("j.kategori='PENDANAAN'")
        rows = _fetch_keluar(q_where + beban_w, q_params)
        total_keluar = total_all = sum(r['total'] for r in rows)
        total_masuk  = 0

    conn.close()
    return render_template('daftar_transaksi.html',
                           rows=rows, tab=tab, beban=beban, q=q,
                           date_from=date_from, date_to=date_to,
                           total_all=total_all,
                           total_masuk=total_masuk, total_keluar=total_keluar)


@app.route('/daftar-transaksi/export')
@login_required
def daftar_transaksi_export():
    if not HAS_XLSX:
        flash('Library openpyxl tidak tersedia.', 'danger')
        return redirect(url_for('daftar_transaksi'))

    conn = db()
    tab       = request.args.get('tab', 'semua')
    beban     = request.args.get('beban', '')
    q         = request.args.get('q', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')

    q_where  = ["j.keterangan LIKE ?"] if q else []
    q_params = [f'%{q}%'] if q else []
    if date_from:
        q_where.append("j.tanggal >= ?"); q_params.append(date_from)
    if date_to:
        q_where.append("j.tanggal <= ?"); q_params.append(date_to)

    _sub_label_sql = """
        CASE
          WHEN j.tipe_tx='PEMASUKAN' THEN 'Penjualan'
          WHEN j.kategori='INVESTASI' THEN 'Investasi Aset'
          WHEN j.kategori='PENDANAAN' THEN 'Penarikan Owner'
          WHEN EXISTS(SELECT 1 FROM detail_jurnal dj JOIN akun ax ON ax.id=dj.akun_id
                      WHERE dj.jurnal_id=j.id AND ax.kode='1130' AND dj.debit>0)
               THEN 'Bahan Baku'
          WHEN EXISTS(SELECT 1 FROM detail_jurnal dj JOIN akun ax ON ax.id=dj.akun_id
                      WHERE dj.jurnal_id=j.id AND ax.kode='6170' AND dj.debit>0)
               THEN 'Pajak'
          ELSE 'Operasional'
        END as sub_label
    """

    def _fetch_masuk(ew=[], ep=[]):
        w = ["j.tipe_tx='PEMASUKAN'"] + ew
        ws = 'WHERE ' + ' AND '.join(w)
        return conn.execute(f"""
            SELECT j.id, j.tanggal, j.keterangan, 'PEMASUKAN' as tipe_tx,
                   (SELECT COALESCE(SUM(dj2.kredit),0)
                    FROM detail_jurnal dj2 JOIN akun a2 ON a2.id=dj2.akun_id
                    WHERE dj2.jurnal_id=j.id AND a2.kode LIKE '4%') as total,
                   {_sub_label_sql}
            FROM jurnal j {ws} ORDER BY j.tanggal, j.id
        """, ep).fetchall()

    def _fetch_keluar(ew=[], ep=[]):
        w = ["j.tipe_tx='PENGELUARAN'"] + ew
        ws = 'WHERE ' + ' AND '.join(w)
        beban_w = list(ew)
        if beban == 'bahan_baku':
            beban_w.append("EXISTS(SELECT 1 FROM detail_jurnal dj JOIN akun a ON a.id=dj.akun_id WHERE dj.jurnal_id=j.id AND a.kode='1130' AND dj.debit>0)")
        elif beban == 'operasional':
            beban_w += ["j.kategori='OPERASIONAL'",
                        "NOT EXISTS(SELECT 1 FROM detail_jurnal dj JOIN akun a ON a.id=dj.akun_id WHERE dj.jurnal_id=j.id AND a.kode IN('1130','6170') AND dj.debit>0)"]
        elif beban == 'investasi':
            beban_w.append("j.kategori='INVESTASI'")
        elif beban == 'pajak':
            beban_w.append("EXISTS(SELECT 1 FROM detail_jurnal dj JOIN akun a ON a.id=dj.akun_id WHERE dj.jurnal_id=j.id AND a.kode='6170' AND dj.debit>0)")
        elif beban == 'penarikan':
            beban_w.append("j.kategori='PENDANAAN'")
        w2 = ["j.tipe_tx='PENGELUARAN'"] + beban_w
        ws2 = 'WHERE ' + ' AND '.join(w2)
        return conn.execute(f"""
            SELECT j.id, j.tanggal, j.keterangan, 'PENGELUARAN' as tipe_tx,
                   COALESCE(SUM(d.debit),0) as total,
                   {_sub_label_sql}
            FROM jurnal j LEFT JOIN detail_jurnal d ON d.jurnal_id=j.id
            {ws2} GROUP BY j.id ORDER BY j.tanggal, j.id
        """, ep).fetchall()

    if tab == 'pemasukan':
        rows = _fetch_masuk(q_where, q_params)
    elif tab == 'pengeluaran':
        rows = _fetch_keluar(q_where, q_params)
    else:
        rp = [dict(r) for r in _fetch_masuk(q_where, q_params)]
        rk = [dict(r) for r in _fetch_keluar(q_where, q_params)]
        rows = sorted(rp + rk, key=lambda r: (r['tanggal'], r['id']))

    conn.close()

    # ── Build Excel ───────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    tab_label = {'semua': 'Semua', 'pemasukan': 'Pemasukan', 'pengeluaran': 'Pengeluaran'}.get(tab, tab)
    ws.title = f'Transaksi {tab_label}'

    # judul
    ws.merge_cells('A1:E1')
    ws['A1'] = f'Daftar Transaksi — {tab_label}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = f'Periode: {date_from or "—"}  s/d  {date_to or "—"}'
    ws['A2'].font = Font(italic=True, size=10, color='666666')
    ws.append([])

    # header
    hdr = ['Tanggal', 'Keterangan', 'Jenis', 'Tipe', 'Nominal (Rp)']
    ws.append(hdr)
    hrow = ws.max_row
    hfill = PatternFill('solid', fgColor='1E293B')
    hfont = Font(bold=True, color='FFFFFF', size=10)
    for col in range(1, 6):
        cell = ws.cell(hrow, col)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    # data
    total_masuk = total_keluar = 0
    for r in rows:
        is_masuk = r['tipe_tx'] == 'PEMASUKAN'
        nominal = r['total']
        if is_masuk:
            total_masuk += nominal
        else:
            total_keluar += nominal
        ws.append([
            r['tanggal'],
            r['keterangan'],
            r['sub_label'],
            'Pemasukan' if is_masuk else 'Pengeluaran',
            nominal if is_masuk else -nominal,
        ])
        drow = ws.max_row
        ws.cell(drow, 5).number_format = '#,##0'
        ws.cell(drow, 5).font = Font(
            color='16A34A' if is_masuk else 'DC2626', bold=True)

    # footer
    ws.append([])
    frow = ws.max_row + 1
    if tab == 'semua':
        ws.cell(frow, 4).value = 'Total Pemasukan'
        ws.cell(frow, 5).value = total_masuk
        ws.cell(frow, 5).number_format = '#,##0'
        ws.cell(frow, 5).font = Font(bold=True, color='16A34A')
        ws.cell(frow+1, 4).value = 'Total Pengeluaran'
        ws.cell(frow+1, 5).value = -total_keluar
        ws.cell(frow+1, 5).number_format = '#,##0'
        ws.cell(frow+1, 5).font = Font(bold=True, color='DC2626')
        ws.cell(frow+2, 4).value = 'Selisih'
        ws.cell(frow+2, 5).value = total_masuk - total_keluar
        ws.cell(frow+2, 5).number_format = '#,##0'
        ws.cell(frow+2, 5).font = Font(bold=True, size=12)
    elif tab == 'pemasukan':
        ws.cell(frow, 4).value = 'Total Pemasukan'
        ws.cell(frow, 5).value = total_masuk
        ws.cell(frow, 5).number_format = '#,##0'
        ws.cell(frow, 5).font = Font(bold=True, color='16A34A')
    else:
        ws.cell(frow, 4).value = 'Total Pengeluaran'
        ws.cell(frow, 5).value = -total_keluar
        ws.cell(frow, 5).number_format = '#,##0'
        ws.cell(frow, 5).font = Font(bold=True, color='DC2626')

    # lebar kolom
    for col, w in zip(range(1, 6), [13, 42, 18, 14, 18]):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname_parts = ['transaksi', tab]
    if date_from: fname_parts.append(date_from)
    if date_to:   fname_parts.append(date_to)
    fname = '_'.join(fname_parts) + '.xlsx'
    return send_file(buf, download_name=fname, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------- TRANSAKSI (Advanced Journal) ----------
@app.route('/transaksi')
@investor_required
def transaksi_list():
    conn = db()
    q = request.args.get('q',''); kat = request.args.get('kat','')
    params=[]; where=[]
    if q: where.append("(j.keterangan LIKE ? OR j.referensi LIKE ?)"); params+=[f'%{q}%',f'%{q}%']
    if kat: where.append("j.kategori=?"); params.append(kat)
    ws = ('WHERE '+' AND '.join(where)) if where else ''
    rows = conn.execute(f"""
        SELECT j.id,j.tanggal,j.keterangan,j.referensi,j.kategori,j.tipe_tx,
               COALESCE(SUM(d.debit),0) as total
        FROM jurnal j LEFT JOIN detail_jurnal d ON d.jurnal_id=j.id
        {ws} GROUP BY j.id ORDER BY j.tanggal DESC, j.id DESC
    """, params).fetchall()
    conn.close()
    return render_template('transaksi.html', rows=rows, q=q, kat=kat)

@app.route('/transaksi/baru', methods=['GET','POST'])
@finance_required
def transaksi_baru_form():
    conn = db()
    if request.method == 'POST':
        tanggal    = request.form['tanggal']
        keterangan = request.form['keterangan']
        referensi  = request.form.get('referensi','')
        kategori   = request.form.get('kategori','OPERASIONAL')
        akun_ids   = request.form.getlist('akun_id[]')
        debits     = request.form.getlist('debit[]')
        kredits    = request.form.getlist('kredit[]')
        total_d = sum(float(x or 0) for x in debits)
        total_k = sum(float(x or 0) for x in kredits)
        if abs(total_d - total_k) > 0.01:
            flash('Total debit dan kredit harus seimbang!','danger')
            akun_list = conn.execute('SELECT * FROM akun ORDER BY kode').fetchall()
            conn.close()
            return render_template('transaksi_form.html', akun_list=akun_list)
        cur = conn.execute(
            'INSERT INTO jurnal(tanggal,keterangan,referensi,kategori,tipe_tx) VALUES(?,?,?,?,?)',
            (tanggal, keterangan, referensi, kategori, 'JURNAL')
        )
        jid = cur.lastrowid
        for i in range(len(akun_ids)):
            if akun_ids[i]:
                d = float(debits[i] or 0); k = float(kredits[i] or 0)
                if d>0 or k>0:
                    conn.execute(
                        'INSERT INTO detail_jurnal(jurnal_id,akun_id,debit,kredit) VALUES(?,?,?,?)',
                        (jid, int(akun_ids[i]), d, k)
                    )
        conn.commit(); conn.close()
        flash('Jurnal berhasil disimpan!','success')
        return redirect(url_for('transaksi_list'))
    akun_list = conn.execute('SELECT * FROM akun ORDER BY kode').fetchall()
    conn.close()
    return render_template('transaksi_form.html', akun_list=akun_list)

@app.route('/transaksi/<int:id>')
@investor_required
def transaksi_detail(id):
    conn = db()
    j = conn.execute('SELECT * FROM jurnal WHERE id=?',(id,)).fetchone()
    if not j:
        flash('Tidak ditemukan','danger')
        return redirect(url_for('transaksi_list'))
    lines = conn.execute("""
        SELECT dj.*,a.kode,a.nama,a.tipe FROM detail_jurnal dj
        JOIN akun a ON a.id=dj.akun_id WHERE dj.jurnal_id=?
    """,(id,)).fetchall()
    conn.close()
    return render_template('transaksi_detail.html', j=j, lines=lines)

@app.route('/transaksi/<int:id>/edit', methods=['GET','POST'])
@finance_required
def transaksi_edit(id):
    conn = db()
    j = conn.execute('SELECT * FROM jurnal WHERE id=?', (id,)).fetchone()
    if not j:
        conn.close()
        flash('Jurnal tidak ditemukan.', 'danger')
        return redirect(url_for('transaksi_list'))

    # Operator: only allowed to edit current-month journals
    if session.get('role') == 'OPERATOR' and not _operator_date_ok(j['tanggal']):
        conn.close()
        flash('Operator hanya bisa mengedit jurnal bulan berjalan.', 'warning')
        return redirect(url_for('transaksi_detail', id=id))

    akun_list = conn.execute('SELECT * FROM akun ORDER BY kode').fetchall()

    if request.method == 'POST':
        tanggal    = request.form['tanggal']
        keterangan = request.form['keterangan']
        referensi  = request.form.get('referensi', '')
        kategori   = request.form.get('kategori', 'OPERASIONAL')
        akun_ids   = request.form.getlist('akun_id[]')
        debits     = request.form.getlist('debit[]')
        kredits    = request.form.getlist('kredit[]')

        # Operator: new date must also be current month
        if not _operator_date_ok(tanggal):
            flash('Operator hanya bisa mengedit jurnal bulan berjalan.', 'warning')
            lines = conn.execute("""
                SELECT dj.*,a.kode,a.nama FROM detail_jurnal dj
                JOIN akun a ON a.id=dj.akun_id WHERE dj.jurnal_id=?
            """, (id,)).fetchall()
            conn.close()
            return render_template('transaksi_form.html', akun_list=akun_list,
                                   edit=j, lines=lines, today=date.today().strftime('%Y-%m-%d'))

        total_d = sum(float(x or 0) for x in debits)
        total_k = sum(float(x or 0) for x in kredits)
        if abs(total_d - total_k) > 0.01:
            flash('Total debit dan kredit harus seimbang!', 'danger')
            lines = conn.execute("""
                SELECT dj.*,a.kode,a.nama FROM detail_jurnal dj
                JOIN akun a ON a.id=dj.akun_id WHERE dj.jurnal_id=?
            """, (id,)).fetchall()
            conn.close()
            return render_template('transaksi_form.html', akun_list=akun_list,
                                   edit=j, lines=lines, today=date.today().strftime('%Y-%m-%d'))

        conn.execute(
            'UPDATE jurnal SET tanggal=?,keterangan=?,referensi=?,kategori=? WHERE id=?',
            (tanggal, keterangan, referensi, kategori, id)
        )
        conn.execute('DELETE FROM detail_jurnal WHERE jurnal_id=?', (id,))
        for i in range(len(akun_ids)):
            if akun_ids[i]:
                d = float(debits[i] or 0); k = float(kredits[i] or 0)
                if d > 0 or k > 0:
                    conn.execute(
                        'INSERT INTO detail_jurnal(jurnal_id,akun_id,debit,kredit) VALUES(?,?,?,?)',
                        (id, int(akun_ids[i]), d, k)
                    )
        add_log(conn, 'Edit jurnal', f"{keterangan} | {tanggal}", 'INPUT')
        conn.commit(); conn.close()
        flash('Jurnal berhasil diupdate!', 'success')
        return redirect(url_for('transaksi_detail', id=id))

    lines = conn.execute("""
        SELECT dj.*,a.kode,a.nama FROM detail_jurnal dj
        JOIN akun a ON a.id=dj.akun_id WHERE dj.jurnal_id=?
    """, (id,)).fetchall()
    conn.close()
    return render_template('transaksi_form.html', akun_list=akun_list,
                           edit=j, lines=lines, today=date.today().strftime('%Y-%m-%d'))


@app.route('/transaksi/<int:id>/hapus', methods=['POST'])
@finance_required
def transaksi_hapus(id):
    conn = db()
    j = conn.execute("SELECT keterangan, tanggal FROM jurnal WHERE id=?", (id,)).fetchone()
    detail = f"{j['keterangan']} | {j['tanggal']}" if j else f"ID {id}"
    conn.execute('DELETE FROM jurnal WHERE id=?',(id,))
    add_log(conn, 'Hapus transaksi', detail, 'INPUT')
    conn.commit(); conn.close()
    flash('Transaksi dihapus.','warning')
    return redirect(url_for('transaksi_list'))


# ---------- KALKULATOR HPP ----------
@app.route('/kalkulator-hpp')
@investor_required
def kalkulator_hpp():
    conn = db()
    produk_list = conn.execute(
        "SELECT id, nama, harga_jual, bahan FROM hpp_produk ORDER BY updated_at DESC"
    ).fetchall()
    conn.close()
    import json as _json
    items = []
    for p in produk_list:
        bahan = _json.loads(p['bahan'] or '[]')
        total_hpp = sum(
            (b.get('harga',0)/b.get('frekuensi',1))*b.get('takaran',0)
            for b in bahan if b.get('frekuensi',0) > 0
        )
        items.append({'id': p['id'], 'nama': p['nama'],
                      'harga_jual': p['harga_jual'], 'total_hpp': round(total_hpp, 2)})
    return render_template('hpp_kalkulator.html', produk_list=items)

@app.route('/api/hpp-produk', methods=['GET'])
@investor_required
def hpp_produk_list():
    import json as _json
    conn = db()
    rows = conn.execute(
        "SELECT id, nama, harga_jual, bahan FROM hpp_produk ORDER BY updated_at DESC"
    ).fetchall()
    conn.close()
    result = []
    for p in rows:
        bahan = _json.loads(p['bahan'] or '[]')
        total_hpp = sum(
            (b.get('harga',0)/b.get('frekuensi',1))*b.get('takaran',0)
            for b in bahan if b.get('frekuensi',0) > 0
        )
        result.append({'id': p['id'], 'nama': p['nama'],
                       'harga_jual': p['harga_jual'], 'total_hpp': round(total_hpp, 2)})
    return jsonify(result)

@app.route('/api/hpp-produk/<int:pid>', methods=['GET'])
@investor_required
def hpp_produk_get(pid):
    conn = db()
    p = conn.execute("SELECT * FROM hpp_produk WHERE id=?", (pid,)).fetchone()
    conn.close()
    if not p:
        return jsonify({'error': 'not found'}), 404
    import json as _json
    return jsonify({'id': p['id'], 'nama': p['nama'],
                    'harga_jual': p['harga_jual'],
                    'bahan': _json.loads(p['bahan'] or '[]')})

@app.route('/api/hpp-produk', methods=['POST'])
@finance_required
def hpp_produk_save():
    import json as _json
    data = request.get_json()
    nama = (data.get('nama') or '').strip()
    if not nama:
        return jsonify({'error': 'nama wajib diisi'}), 400
    harga_jual = float(data.get('harga_jual') or 0)
    bahan = _json.dumps(data.get('bahan') or [], ensure_ascii=False)
    conn = db()
    cur = conn.execute(
        "INSERT INTO hpp_produk(nama, harga_jual, bahan, updated_at) VALUES(?,?,?,date('now'))",
        (nama, harga_jual, bahan)
    )
    pid = cur.lastrowid
    conn.commit(); conn.close()
    return jsonify({'id': pid, 'nama': nama})

@app.route('/api/hpp-produk/<int:pid>', methods=['PUT'])
@finance_required
def hpp_produk_update(pid):
    import json as _json
    data = request.get_json()
    nama = (data.get('nama') or '').strip()
    if not nama:
        return jsonify({'error': 'nama wajib diisi'}), 400
    harga_jual = float(data.get('harga_jual') or 0)
    bahan = _json.dumps(data.get('bahan') or [], ensure_ascii=False)
    conn = db()
    conn.execute(
        "UPDATE hpp_produk SET nama=?, harga_jual=?, bahan=?, updated_at=date('now') WHERE id=?",
        (nama, harga_jual, bahan, pid)
    )
    conn.commit(); conn.close()
    return jsonify({'id': pid, 'nama': nama})

@app.route('/api/hpp-produk/<int:pid>', methods=['DELETE'])
@finance_required
def hpp_produk_delete(pid):
    conn = db()
    conn.execute("DELETE FROM hpp_produk WHERE id=?", (pid,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})


# ---------- BACKUP & RESTORE ----------
@app.route('/backup')
@finance_required
def backup():
    import shutil, tempfile
    today = date.today().strftime('%Y%m%d')
    tmp = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
    tmp.close()
    src_conn = sqlite3.connect(DB)
    dst_conn = sqlite3.connect(tmp.name)
    src_conn.backup(dst_conn)
    src_conn.close()
    dst_conn.close()
    conn2 = db()
    add_log(conn2, 'Backup database', f"finansial_backup_{today}.db", 'MODAL')
    conn2.commit(); conn2.close()
    return send_file(tmp.name, as_attachment=True,
                     download_name=f'finansial_backup_{today}.db',
                     mimetype='application/octet-stream')

@app.route('/restore', methods=['POST'])
@finance_required
def restore():
    f = request.files.get('backup_file')
    if not f or not f.filename.endswith('.db'):
        flash('File tidak valid. Gunakan file .db hasil backup.', 'danger')
        return redirect('/settings')
    import tempfile, shutil
    tmp = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
    f.save(tmp.name)
    tmp.close()
    # Validate: try opening and checking tables exist
    try:
        test = sqlite3.connect(tmp.name)
        tables = {r[0] for r in test.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        test.close()
        required = {'jurnal', 'detail_jurnal', 'akun', 'users'}
        if not required.issubset(tables):
            raise ValueError('Bukan file database FinansialApp yang valid.')
    except Exception as e:
        os.unlink(tmp.name)
        flash(f'Restore gagal: {e}', 'danger')
        return redirect('/settings')
    # Replace database
    shutil.copy2(tmp.name, DB)
    os.unlink(tmp.name)
    init_db()  # apply any missing migrations to the restored database
    conn2 = db()
    add_log(conn2, 'Restore database', f"File: {f.filename}", 'MODAL')
    conn2.commit(); conn2.close()
    flash('Database berhasil direstore. Silakan login ulang jika diperlukan.', 'success')
    return redirect('/settings')


# ---------- LOG AKTIVITAS ----------
@app.route('/log')
@finance_required
def log_aktivitas():
    import math
    conn = db()
    page     = int(request.args.get('page', 1))
    per      = 50
    q        = request.args.get('q', '').strip()
    kategori = request.args.get('kategori', '').strip()
    where, params = [], []
    if q:
        where.append("(detail LIKE ? OR username LIKE ? OR aksi LIKE ?)")
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if kategori:
        where.append("kategori=?")
        params.append(kategori)
    where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''
    total = conn.execute(f"SELECT COUNT(*) FROM log_aktivitas {where_sql}", params).fetchone()[0]
    logs  = conn.execute(
        f"SELECT * FROM log_aktivitas {where_sql} ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [per, (page-1)*per]
    ).fetchall()
    # counts per category for stats row
    counts = {}
    for row in conn.execute("SELECT kategori, COUNT(*) as c FROM log_aktivitas GROUP BY kategori"):
        counts[row['kategori']] = row['c']
    conn.close()
    total_pages = max(1, math.ceil(total / per))
    return render_template('log.html',
        logs=logs, page=page, total_pages=total_pages,
        total=total, q=q, kategori=kategori, counts=counts)


# ---------- SETTINGS ----------
@app.route('/settings', methods=['GET','POST'])
@admin_required
def settings():
    conn = db()
    if request.method == 'POST':
        action = request.form.get('action','')
        if action == 'save_settings':
            for key in ['nama_usaha']:
                val = request.form.get(key,'')
                conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (key,val))
            conn.commit()
            flash('Pengaturan disimpan!','success')
        elif action == 'add_user':
            username = request.form.get('username','').strip()
            password = request.form.get('password','')
            nama     = request.form.get('nama','')
            role     = request.form.get('role','VIEWER')
            ph = sha256(password.encode()).hexdigest()
            try:
                conn.execute("INSERT INTO users(username,password_hash,nama,role) VALUES(?,?,?,?)",
                             (username, ph, nama, role))
                add_log(conn, 'Tambah user', f"Username: {username} | Role: {role}", 'ADMIN')
                conn.commit()
                flash(f'User {username} berhasil ditambahkan!','success')
            except sqlite3.IntegrityError:
                flash('Username sudah digunakan!','danger')
        elif action == 'toggle_user':
            uid = int(request.form.get('user_id',0))
            u = conn.execute("SELECT username, aktif FROM users WHERE id=?", (uid,)).fetchone()
            conn.execute("UPDATE users SET aktif=CASE WHEN aktif=1 THEN 0 ELSE 1 END WHERE id=? AND username!='admin'", (uid,))
            if u:
                status_baru = 'dinonaktifkan' if u['aktif'] else 'diaktifkan'
                add_log(conn, 'Toggle status user', f"Username: {u['username']} → {status_baru}", 'ADMIN')
            conn.commit()
            flash('Status user diperbarui.','info')
        elif action == 'reset_password':
            uid = int(request.form.get('user_id',0))
            new_pw = request.form.get('new_password','')
            if new_pw:
                ph = sha256(new_pw.encode()).hexdigest()
                conn.execute("UPDATE users SET password_hash=? WHERE id=?", (ph, uid))
                u = conn.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
                add_log(conn, 'Reset password user', f"Username: {u['username'] if u else uid}", 'PASSWORD')
                conn.commit()
                flash('Password berhasil direset.','success')
        elif action == 'delete_user':
            uid = int(request.form.get('user_id',0))
            u = conn.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
            conn.execute("DELETE FROM users WHERE id=? AND username!='admin'", (uid,))
            add_log(conn, 'Hapus user', f"Username: {u['username'] if u else uid}", 'ADMIN')
            conn.commit()
            flash('User dihapus.','warning')
        elif action == 'change_role':
            uid      = int(request.form.get('user_id', 0))
            new_role = request.form.get('new_role', '').strip()
            if new_role in ('ADMIN', 'FINANCE', 'INVESTOR', 'OPERATOR') and uid:
                u = conn.execute("SELECT username, role FROM users WHERE id=?", (uid,)).fetchone()
                conn.execute("UPDATE users SET role=? WHERE id=? AND username!='admin'", (new_role, uid))
                if u:
                    add_log(conn, 'Ubah role user', f"Username: {u['username']} | {u['role']} → {new_role}", 'ADMIN')
                conn.commit()
                flash(f'Role berhasil diubah ke {new_role}.', 'success')
        elif action == 'modal':
            nominal = float(request.form.get('modal_nominal',0) or 0)
            tanggal = request.form.get('modal_tanggal', str(date.today()))
            akun_kas_kode = request.form.get('akun_kas','1100')
            catatan = (request.form.get('modal_catatan','') or '').strip()
            if nominal > 0:
                ket = 'Setoran Modal'
                if catatan:
                    ket = f"{ket} - {catatan}"
                insert_jurnal(conn, tanggal, ket, 'PENDANAAN', 'JURNAL', [
                    (akun_kas_kode, nominal, 0), ('3100', 0, nominal)
                ])
                add_log(conn, 'Setoran modal', f"Rp {nominal:,.0f} | {tanggal}", 'MODAL')
                conn.commit()
                flash('Setoran modal berhasil dicatat!','success')
            else:
                flash('Nominal modal harus lebih besar dari 0.', 'warning')
        elif action == 'hapus_modal':
            jurnal_id = int(request.form.get('jurnal_id', 0) or 0)
            if jurnal_id:
                row = conn.execute(
                    "SELECT keterangan, tanggal FROM jurnal WHERE id=? AND kategori='PENDANAAN'",
                    (jurnal_id,)
                ).fetchone()
                if row and (row['keterangan'] == 'Setoran Modal' or row['keterangan'].startswith('Setoran Modal')):
                    conn.execute("DELETE FROM jurnal WHERE id=?", (jurnal_id,))
                    add_log(conn, 'Hapus setoran modal', f"Jurnal #{jurnal_id} | {row['tanggal']}", 'MODAL')
                    conn.commit()
                    flash('Setoran modal dihapus.', 'warning')
                else:
                    flash('Jurnal tidak ditemukan atau bukan setoran modal.', 'danger')
        elif action == 'add_rekening':
            rek_nama = request.form.get('rek_nama','').strip()
            rek_no   = request.form.get('rek_no','').strip()
            rek_jenis= request.form.get('rek_jenis','Bank')
            if rek_nama:
                last = conn.execute(
                    "SELECT kode FROM akun WHERE kode GLOB '11[0-9][0-9]' ORDER BY kode DESC LIMIT 1"
                ).fetchone()
                next_num = int(last['kode']) + 1 if last else 1112
                if next_num > 1199:
                    flash('Maksimum 100 rekening kas/bank.','danger')
                else:
                    nama_akun = f"{rek_jenis} {rek_nama}"
                    conn.execute(
                        "INSERT INTO akun(kode,nama,tipe,subtipe,saldo_normal,is_rekening,no_rekening) VALUES(?,?,?,?,?,?,?)",
                        (str(next_num), nama_akun, 'ASET', 'Aset Lancar', 'DEBIT', 1, rek_no or None)
                    )
                    conn.commit()
                    flash(f'Rekening {nama_akun} berhasil ditambahkan!','success')
        elif action == 'hapus_rekening':
            aid = int(request.form.get('akun_id',0))
            a = conn.execute("SELECT * FROM akun WHERE id=?", (aid,)).fetchone()
            if a and a['kode'] in ('1100','1110'):
                flash('Rekening Kas dan Bank utama tidak bisa dihapus.','danger')
            elif a:
                used = conn.execute("SELECT COUNT(*) FROM detail_jurnal WHERE akun_id=?", (aid,)).fetchone()[0]
                if used > 0:
                    flash('Rekening sudah digunakan dalam transaksi, tidak bisa dihapus.','danger')
                else:
                    conn.execute("DELETE FROM akun WHERE id=?", (aid,))
                    conn.commit()
                    flash('Rekening dihapus.','warning')
        elif action == 'edit_rekening':
            aid      = int(request.form.get('akun_id', 0))
            rek_nama = request.form.get('rek_nama', '').strip()
            rek_no   = request.form.get('rek_no', '').strip()
            if aid and rek_nama:
                conn.execute("UPDATE akun SET nama=?, no_rekening=? WHERE id=? AND is_rekening=1",
                             (rek_nama, rek_no or None, aid))
                conn.commit()
                flash('Rekening berhasil diperbarui!', 'success')
        elif action == 'toggle_rekening':
            aid = int(request.form.get('akun_id',0))
            a = conn.execute("SELECT * FROM akun WHERE id=?", (aid,)).fetchone()
            if a and a['kode'] not in ('1100','1110'):
                new_flag = 0 if a['is_rekening'] else 1
                conn.execute("UPDATE akun SET is_rekening=? WHERE id=?", (new_flag, aid))
                conn.commit()
                flash('Status rekening diperbarui.','info')
        conn.close()
        return redirect(url_for('settings'))

    s_nama_usaha   = get_setting(conn, 'nama_usaha','Usaha Saya')
    users = conn.execute("SELECT * FROM users ORDER BY role, username").fetchall()
    akun_kas = conn.execute("SELECT * FROM akun WHERE is_rekening=1 ORDER BY kode").fetchall()
    rekening_list = get_rekening_saldo(conn)
    # Riwayat setoran modal: ambil semua jurnal PENDANAAN dgn keterangan 'Setoran Modal*'
    riwayat_modal = conn.execute("""
        SELECT j.id, j.tanggal, j.keterangan,
               COALESCE(SUM(d.kredit),0) as jumlah,
               (SELECT a.nama FROM detail_jurnal dd
                JOIN akun a ON a.id=dd.akun_id
                WHERE dd.jurnal_id=j.id AND dd.debit>0 LIMIT 1) as akun_kas_nama
        FROM jurnal j
        LEFT JOIN detail_jurnal d ON d.jurnal_id=j.id
        LEFT JOIN akun a2 ON a2.id=d.akun_id
        WHERE j.kategori='PENDANAAN'
          AND (j.keterangan='Setoran Modal'
               OR j.keterangan LIKE 'Setoran Modal -%'
               OR j.keterangan='Setoran Modal Awal'
               OR j.keterangan LIKE 'Setoran Modal Awal%')
          AND a2.kode='3100'
        GROUP BY j.id
        ORDER BY j.tanggal DESC, j.id DESC
    """).fetchall()
    total_modal = sum(r['jumlah'] for r in riwayat_modal)
    conn.close()
    return render_template('settings.html',
        s_nama_usaha=s_nama_usaha, users=users, akun_kas=akun_kas,
        rekening_list=rekening_list,
        riwayat_modal=riwayat_modal, total_modal=total_modal)


# ---------- ASET TETAP ----------
@app.route('/aset-tetap')
@investor_required
def aset_tetap_list():
    conn = db()
    today = date.today()
    raw = conn.execute("SELECT * FROM aset_tetap ORDER BY tanggal_beli DESC").fetchall()
    aset_list = []
    for a in raw:
        tgl_beli = datetime.strptime(str(a['tanggal_beli'])[:10], '%Y-%m-%d').date()
        months_elapsed = (today.year - tgl_beli.year) * 12 + (today.month - tgl_beli.month) + 1
        bulan_real = min(months_elapsed, a['masa_pakai'])
        peny_real  = round(bulan_real * (a['penyusutan_bulan'] or 0), 0)
        bulan_ctt  = a['bulan_penyusutan_dicatat'] or 0
        row = dict(a)
        row['bulan_real']  = bulan_real
        row['peny_real']   = peny_real
        row['bulan_ctt']   = bulan_ctt
        aset_list.append(row)
    conn.close()
    return render_template('aset_tetap.html', aset_list=aset_list, today=today)


# ---------- AKUN ----------
@app.route('/akun')
@investor_required
def akun_list():
    conn = db()
    rows = conn.execute('SELECT * FROM akun ORDER BY kode').fetchall()
    conn.close()
    return render_template('akun.html', rows=rows)

@app.route('/akun/baru', methods=['POST'])
@finance_required
def akun_baru():
    conn = db()
    try:
        conn.execute(
            'INSERT INTO akun(kode,nama,tipe,subtipe,saldo_normal) VALUES(?,?,?,?,?)',
            (request.form['kode'], request.form['nama'], request.form['tipe'],
             request.form.get('subtipe',''), request.form['saldo_normal'])
        )
        conn.commit(); flash('Akun ditambahkan!','success')
    except sqlite3.IntegrityError:
        flash('Kode akun sudah ada!','danger')
    finally:
        conn.close()
    return redirect(url_for('akun_list'))

@app.route('/akun/<int:id>/hapus', methods=['POST'])
@finance_required
def akun_hapus(id):
    conn = db()
    used = conn.execute('SELECT COUNT(*) FROM detail_jurnal WHERE akun_id=?',(id,)).fetchone()[0]
    if used > 0:
        flash('Akun sudah digunakan, tidak bisa dihapus!','danger')
    else:
        conn.execute('DELETE FROM akun WHERE id=?',(id,))
        conn.commit(); flash('Akun dihapus.','warning')
    conn.close()
    return redirect(url_for('akun_list'))


# ---------- EXPORT ----------
@app.route('/export/dashboard')
@investor_required
def export_dashboard():
    today = date.today()
    sd = request.args.get('sd', today.replace(day=1).strftime('%Y-%m-%d'))
    ed = request.args.get('ed', today.strftime('%Y-%m-%d'))
    conn = db()
    pnl = calc_profitability(conn, sd, ed)
    cf  = calc_cashflow(conn, sd, ed)
    conn.close()

    if HAS_XLSX:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dashboard"
        hdr = Font(bold=True)
        ws.append(["Laporan Keuangan", f"{sd} s/d {ed}"])
        ws.append([])
        ws.append(["CASHFLOW"])
        ws.append(["Uang Masuk", cf['masuk']])
        ws.append(["Uang Keluar", cf['keluar']])
        ws.append(["Saldo", cf['saldo']])
        ws.append([])
        ws.append(["PROFITABILITAS", "Nominal", "%"])
        items = [
            ("Pendapatan", pnl['rev'], 100),
            ("HPP", pnl['hpp'], pnl['pct_hpp']),
            ("Laba Kotor", pnl['laba_kotor'], pnl['pct_laba_kotor']),
            ("Biaya Operasional", pnl['op_exp'], pnl['pct_op']),
            ("Laba Operasional", pnl['laba_op'], pnl['pct_laba_op']),
            ("Penyusutan", pnl['depr'], pnl['pct_depr']),
            ("EBIT", pnl['ebit'], pnl['pct_ebit']),
            ("Pajak", pnl['tax'], pnl['pct_tax']),
            ("Laba Bersih", pnl['laba_bersih'], pnl['pct_laba_bersih']),
            ("Penarikan Owner", pnl['prive'], pnl['pct_prive']),
            ("Laba Ditahan", pnl['laba_tahan'], pnl['pct_laba_tahan']),
        ]
        for row in items:
            ws.append(list(row))
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, download_name=f"dashboard_{sd}_{ed}.xlsx",
                         as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # CSV fallback
    lines = [f"Laporan Keuangan {sd} s/d {ed}"]
    lines += ["","CASHFLOW",f"Uang Masuk,{cf['masuk']}",f"Uang Keluar,{cf['keluar']}",f"Saldo,{cf['saldo']}"]
    lines += ["","PROFITABILITAS,Nominal,%"]
    for name, val, pct in [("Pendapatan",pnl['rev'],100),("HPP",pnl['hpp'],pnl['pct_hpp']),
                            ("Laba Kotor",pnl['laba_kotor'],pnl['pct_laba_kotor']),
                            ("Biaya Operasional",pnl['op_exp'],pnl['pct_op']),
                            ("Laba Operasional",pnl['laba_op'],pnl['pct_laba_op']),
                            ("Penyusutan",pnl['depr'],pnl['pct_depr']),
                            ("EBIT",pnl['ebit'],pnl['pct_ebit']),
                            ("Pajak",pnl['tax'],pnl['pct_tax']),
                            ("Laba Bersih",pnl['laba_bersih'],pnl['pct_laba_bersih']),
                            ("Penarikan Owner",pnl['prive'],pnl['pct_prive']),
                            ("Laba Ditahan",pnl['laba_tahan'],pnl['pct_laba_tahan'])]:
        lines.append(f"{name},{val},{pct}%")
    csv_data = '\n'.join(lines)
    return Response(csv_data, mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment; filename=dashboard_{sd}_{ed}.csv'})

@app.route('/export/piutang')
@investor_required
def export_piutang():
    conn = db()
    rows = conn.execute("SELECT * FROM piutang ORDER BY status, jatuh_tempo").fetchall()
    conn.close()
    if HAS_XLSX:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Piutang"
        ws.append(["Tanggal","Jatuh Tempo","Pelanggan","Keterangan","Jumlah","Terbayar","Sisa","Status"])
        for r in rows:
            ws.append([r['tanggal'],r['jatuh_tempo'],r['pelanggan'],r['keterangan'],
                       r['jumlah'],r['terbayar'],r['jumlah']-r['terbayar'],r['status']])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, download_name='piutang.xlsx', as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    lines = ["Tanggal,Jatuh Tempo,Pelanggan,Keterangan,Jumlah,Terbayar,Sisa,Status"]
    for r in rows:
        lines.append(f"{r['tanggal']},{r['jatuh_tempo']},{r['pelanggan']},{r['keterangan']},"
                     f"{r['jumlah']},{r['terbayar']},{r['jumlah']-r['terbayar']},{r['status']}")
    return Response('\n'.join(lines), mimetype='text/csv',
                    headers={'Content-Disposition':'attachment; filename=piutang.csv'})

@app.route('/export/hutang')
@investor_required
def export_hutang():
    conn = db()
    rows = conn.execute("SELECT * FROM hutang ORDER BY status, jatuh_tempo").fetchall()
    conn.close()
    if HAS_XLSX:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Hutang"
        ws.append(["Tanggal","Jatuh Tempo","Pemasok","Keterangan","Jumlah","Terbayar","Sisa","Status"])
        for r in rows:
            ws.append([r['tanggal'],r['jatuh_tempo'],r['pemasok'],r['keterangan'],
                       r['jumlah'],r['terbayar'],r['jumlah']-r['terbayar'],r['status']])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, download_name='hutang.xlsx', as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    lines = ["Tanggal,Jatuh Tempo,Pemasok,Keterangan,Jumlah,Terbayar,Sisa,Status"]
    for r in rows:
        lines.append(f"{r['tanggal']},{r['jatuh_tempo']},{r['pemasok']},{r['keterangan']},"
                     f"{r['jumlah']},{r['terbayar']},{r['jumlah']-r['terbayar']},{r['status']}")
    return Response('\n'.join(lines), mimetype='text/csv',
                    headers={'Content-Disposition':'attachment; filename=hutang.csv'})


@app.route('/export/laporan-keuangan')
@investor_required
def export_laporan_keuangan():
    today = date.today()
    sd = request.args.get('sd', today.replace(day=1).strftime('%Y-%m-%d'))
    ed = request.args.get('ed', today.strftime('%Y-%m-%d'))

    conn = db()

    # ── Data Laba Rugi ──────────────────────────────────────────────────────
    pnl = calc_profitability(conn, sd, ed)
    lr_rows = conn.execute("""
        SELECT a.kode, a.nama, a.tipe, a.subtipe,
               COALESCE(SUM(CASE WHEN a.saldo_normal='DEBIT' THEN d.debit-d.kredit
                                 ELSE d.kredit-d.debit END),0) as saldo
        FROM akun a
        LEFT JOIN (
            SELECT dj.akun_id, dj.debit, dj.kredit
            FROM detail_jurnal dj
            JOIN jurnal j ON j.id = dj.jurnal_id AND j.tanggal >= ? AND j.tanggal <= ?
        ) d ON d.akun_id = a.id
        WHERE a.tipe IN ('PENDAPATAN','BEBAN')
        GROUP BY a.id ORDER BY a.kode
    """, (sd, ed)).fetchall()
    pendapatan = {}; beban = {}
    for r in lr_rows:
        item = dict(kode=r['kode'], nama=r['nama'], saldo=r['saldo'])
        if r['tipe'] == 'PENDAPATAN':
            pendapatan.setdefault(r['subtipe'] or 'Pendapatan', []).append(item)
        else:
            beban.setdefault(r['subtipe'] or 'Beban', []).append(item)

    # ── Data Arus Kas ───────────────────────────────────────────────────────
    kas_ids = get_rekening_ids(conn)
    saldo_awal = 0; operasional = []; investasi = []; pendanaan_list = []
    if kas_ids:
        ph = ','.join('?' * len(kas_ids))
        saldo_awal = conn.execute(f"""
            SELECT COALESCE(SUM(d.debit-d.kredit),0)
            FROM detail_jurnal d JOIN jurnal j ON j.id=d.jurnal_id
            WHERE d.akun_id IN ({ph}) AND j.tanggal<?
        """, kas_ids + [sd]).fetchone()[0]
        flows = conn.execute(f"""
            SELECT j.id, j.tanggal, j.keterangan, j.kategori,
                   SUM(CASE WHEN d.akun_id IN ({ph}) THEN d.debit-d.kredit ELSE 0 END) as net_kas
            FROM jurnal j JOIN detail_jurnal d ON d.jurnal_id=j.id
            WHERE j.tanggal>=? AND j.tanggal<=?
            GROUP BY j.id HAVING net_kas!=0 ORDER BY j.tanggal, j.id
        """, kas_ids + [sd, ed]).fetchall()
        operasional   = [r for r in flows if r['kategori'] == 'OPERASIONAL']
        investasi     = [r for r in flows if r['kategori'] == 'INVESTASI']
        pendanaan_list = [r for r in flows if r['kategori'] == 'PENDANAAN']
    total_op   = sum(r['net_kas'] for r in operasional)
    total_inv  = sum(r['net_kas'] for r in investasi)
    total_pend = sum(r['net_kas'] for r in pendanaan_list)
    saldo_akhir = saldo_awal + total_op + total_inv + total_pend

    # ── Data Neraca (per ed) ────────────────────────────────────────────────
    nr_rows = conn.execute("""
        SELECT a.kode, a.nama, a.tipe, a.subtipe, a.saldo_normal,
               COALESCE(SUM(d.debit),0) as td, COALESCE(SUM(d.kredit),0) as tk
        FROM akun a
        LEFT JOIN (
            SELECT dj.akun_id, dj.debit, dj.kredit
            FROM detail_jurnal dj
            JOIN jurnal j ON j.id = dj.jurnal_id AND j.tanggal <= ?
        ) d ON d.akun_id = a.id
        GROUP BY a.id ORDER BY a.kode
    """, (ed,)).fetchall()
    aset = {}; liab = {}; ekuitas_list = []
    total_aset = total_liab = total_ekuitas = 0
    for r in nr_rows:
        saldo = (r['td'] - r['tk']) if r['saldo_normal'] == 'DEBIT' else (r['tk'] - r['td'])
        item = dict(kode=r['kode'], nama=r['nama'], saldo=saldo)
        if r['tipe'] == 'ASET':
            aset_saldo = -saldo if r['saldo_normal'] == 'KREDIT' else saldo
            aset.setdefault(r['subtipe'] or 'Aset Lancar', []).append(
                dict(kode=r['kode'], nama=r['nama'], saldo=aset_saldo))
            total_aset += aset_saldo
        elif r['tipe'] == 'LIABILITAS':
            liab.setdefault(r['subtipe'] or 'Liabilitas Lancar', []).append(item)
            total_liab += saldo
        elif r['tipe'] == 'EKUITAS':
            eks = -saldo if r['saldo_normal'] == 'DEBIT' else saldo
            ekuitas_list.append(dict(kode=r['kode'], nama=r['nama'], saldo=eks))
            total_ekuitas += eks
    fiscal_start = f"{ed[:4]}-01-01"
    laba_tahun   = calc_profitability(conn, fiscal_start, ed)['laba_bersih']
    laba_all     = calc_profitability(conn, '2000-01-01', ed)['laba_bersih']
    laba_ditahan = laba_all - laba_tahun
    if abs(laba_ditahan) > 0.01:
        ekuitas_list.append(dict(kode='-', nama='Laba Ditahan (Akumulasi)', saldo=laba_ditahan))
        total_ekuitas += laba_ditahan
    ekuitas_list.append(dict(kode='-', nama=f'Laba Bersih Tahun {ed[:4]}', saldo=laba_tahun))
    total_ekuitas += laba_tahun

    conn.close()

    if not HAS_XLSX:
        flash('Library openpyxl tidak tersedia. Install dengan: pip install openpyxl', 'danger')
        return redirect(request.referrer or url_for('laba_rugi'))

    wb = openpyxl.Workbook()

    # ── Style helpers ───────────────────────────────────────────────────────
    def _hdr1(ws, text, row):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=14, color='FFFFFF')
        c.fill = PatternFill(patternType='solid', fgColor='1E293B')
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 22

    def _hdr2(ws, text, row):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill(patternType='solid', fgColor='334155')
        ws.row_dimensions[row].height = 16

    def _subtipe_row(ws, text, row, ncols=3):
        c = ws.cell(row=row, column=1, value=text.upper())
        c.font = Font(bold=True, size=9, color='475569')
        c.fill = PatternFill(patternType='solid', fgColor='F1F5F9')
        for col in range(2, ncols + 1):
            ws.cell(row=row, column=col).fill = PatternFill(patternType='solid', fgColor='F1F5F9')

    def _total_row(ws, label, value, row, ncols=3, color='E2E8F0'):
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill(patternType='solid', fgColor=color)
        v = ws.cell(row=row, column=ncols, value=value)
        v.font = Font(bold=True, size=10)
        v.number_format = '#,##0'
        v.fill = PatternFill(patternType='solid', fgColor=color)
        for col in range(2, ncols):
            ws.cell(row=row, column=col).fill = PatternFill(patternType='solid', fgColor=color)

    def _data_row(ws, kode, nama, value, row):
        ws.cell(row=row, column=1, value=kode).font = Font(color='94A3B8', size=9)
        ws.cell(row=row, column=2, value=nama).font = Font(size=10)
        v = ws.cell(row=row, column=3, value=value)
        v.number_format = '#,##0'
        v.font = Font(size=10, color='DC2626' if value < 0 else '000000')

    def _info_row(ws, label, value, row, bold=False):
        c = ws.cell(row=row, column=1, value=label)
        v = ws.cell(row=row, column=3, value=value)
        if bold:
            c.font = Font(bold=True, size=10)
            v.font = Font(bold=True, size=10)
        v.number_format = '#,##0'

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1: LABA RUGI
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Laba Rugi'
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 36
    ws1.column_dimensions['C'].width = 18

    r = 1
    _hdr1(ws1, 'LAPORAN LABA RUGI', r); r += 1
    ws1.cell(row=r, column=1, value=f'Periode: {sd}  s/d  {ed}').font = Font(italic=True, size=9, color='64748B'); r += 2

    # Pendapatan
    _hdr2(ws1, 'PENDAPATAN', r); r += 1
    for subtipe, items in pendapatan.items():
        _subtipe_row(ws1, subtipe, r); r += 1
        sub = 0
        for it in items:
            _data_row(ws1, it['kode'], it['nama'], it['saldo'], r); r += 1
            sub += it['saldo']
        _total_row(ws1, f'  Total {subtipe}', sub, r); r += 1
    _total_row(ws1, 'TOTAL PENDAPATAN', pnl['rev'], r, color='DBEAFE'); r += 2

    # Beban
    _hdr2(ws1, 'BEBAN', r); r += 1
    for subtipe, items in beban.items():
        _subtipe_row(ws1, subtipe, r); r += 1
        sub = 0
        for it in items:
            _data_row(ws1, it['kode'], it['nama'], it['saldo'], r); r += 1
            sub += it['saldo']
        _total_row(ws1, f'  Total {subtipe}', sub, r); r += 1
    _total_row(ws1, 'TOTAL BEBAN', pnl['hpp'] + pnl['op_exp'] + pnl['depr'] + pnl['interest'] + pnl['tax'], r, color='FEE2E2'); r += 2

    # Ringkasan P&L
    _hdr2(ws1, 'RINGKASAN', r); r += 1
    summary = [
        ('Pendapatan',          pnl['rev'],          False),
        ('HPP / Beban Pokok',   -pnl['hpp'],          False),
        ('Laba Kotor',          pnl['laba_kotor'],    True),
        ('Biaya Operasional',   -pnl['op_exp'],       False),
        ('Laba Operasional',    pnl['laba_op'],       True),
        ('Penyusutan',          -pnl['depr'],         False),
        ('EBIT',                pnl['ebit'],          True),
        ('Pajak',               -pnl['tax'],          False),
        ('Laba Bersih',         pnl['laba_bersih'],   True),
        ('Penarikan Owner',     -pnl['prive'],        False),
        ('Laba Ditahan',        pnl['laba_tahan'],    True),
    ]
    for label, val, bold in summary:
        _info_row(ws1, label, val, r, bold=bold); r += 1

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 2: ARUS KAS
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Arus Kas')
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 18

    r = 1
    _hdr1(ws2, 'LAPORAN ARUS KAS', r); r += 1
    ws2.cell(row=r, column=1, value=f'Periode: {sd}  s/d  {ed}').font = Font(italic=True, size=9, color='64748B'); r += 2

    ws2.cell(row=r, column=1, value='Saldo Kas Awal').font = Font(bold=True, size=10)
    v = ws2.cell(row=r, column=3, value=saldo_awal)
    v.number_format = '#,##0'; v.font = Font(bold=True, size=10)
    r += 2

    def _ak_section(ws, title, flows, total, row, hdr_color):
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill(patternType='solid', fgColor=hdr_color)
        ws.cell(row=row, column=2).fill = PatternFill(patternType='solid', fgColor=hdr_color)
        ws.cell(row=row, column=3).fill = PatternFill(patternType='solid', fgColor=hdr_color)
        row += 1
        for flow in flows:
            ws.cell(row=row, column=1, value=str(flow['tanggal'])[:10]).font = Font(size=9, color='94A3B8')
            ws.cell(row=row, column=2, value=flow['keterangan']).font = Font(size=9)
            v = ws.cell(row=row, column=3, value=flow['net_kas'])
            v.number_format = '#,##0'
            v.font = Font(size=9, color='15803D' if flow['net_kas'] >= 0 else 'DC2626')
            row += 1
        _total_row(ws, f'  Total {title}', total, row)
        row += 2
        return row

    r = _ak_section(ws2, 'Aktivitas Operasional', operasional, total_op, r, '0369A1')
    r = _ak_section(ws2, 'Aktivitas Investasi',   investasi,   total_inv, r, '0369A1')
    r = _ak_section(ws2, 'Aktivitas Pendanaan',   pendanaan_list, total_pend, r, '0369A1')

    ws2.cell(row=r, column=1, value='Saldo Kas Akhir').font = Font(bold=True, size=11)
    v = ws2.cell(row=r, column=3, value=saldo_akhir)
    v.number_format = '#,##0'
    v.font = Font(bold=True, size=11, color='1D4ED8')

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 3: NERACA
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Neraca')
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 36
    ws3.column_dimensions['C'].width = 18

    try:
        r = 1
        _hdr1(ws3, 'NERACA (BALANCE SHEET)', r); r += 1
        ws3.cell(row=r, column=1, value=f'Per tanggal: {ed}').font = Font(italic=True, size=9, color='64748B'); r += 2

        for tipe_label, sections, grand_total, hdr_color, total_label in [
            ('ASET',       aset, total_aset, '1D4ED8', 'TOTAL ASET'),
            ('LIABILITAS', liab, total_liab, 'DC2626', 'TOTAL LIABILITAS'),
        ]:
            c = ws3.cell(row=r, column=1, value=tipe_label)
            c.font = Font(bold=True, size=11, color='FFFFFF')
            c.fill = PatternFill(patternType='solid', fgColor=hdr_color)
            ws3.cell(row=r, column=2).fill = PatternFill(patternType='solid', fgColor=hdr_color)
            ws3.cell(row=r, column=3).fill = PatternFill(patternType='solid', fgColor=hdr_color)
            r += 1
            for subtipe, items in sections.items():
                _subtipe_row(ws3, subtipe, r); r += 1
                sub = 0
                for it in items:
                    _data_row(ws3, it['kode'], it['nama'], it['saldo'], r); r += 1
                    sub += it['saldo']
                _total_row(ws3, f'  Total {subtipe}', sub, r); r += 1
            tot_color = 'BFDBFE' if hdr_color == '1D4ED8' else 'FECACA'
            _total_row(ws3, total_label, grand_total, r, color=tot_color); r += 2

        # Ekuitas
        c = ws3.cell(row=r, column=1, value='EKUITAS')
        c.font = Font(bold=True, size=11, color='FFFFFF')
        c.fill = PatternFill(patternType='solid', fgColor='15803D')
        ws3.cell(row=r, column=2).fill = PatternFill(patternType='solid', fgColor='15803D')
        ws3.cell(row=r, column=3).fill = PatternFill(patternType='solid', fgColor='15803D')
        r += 1
        for it in ekuitas_list:
            _data_row(ws3, it['kode'], it['nama'], it['saldo'], r); r += 1
        _total_row(ws3, 'TOTAL EKUITAS', total_ekuitas, r, color='BBF7D0'); r += 2

        # Cek seimbang
        selisih = total_aset - (total_liab + total_ekuitas)
        _total_row(ws3, 'TOTAL LIABILITAS + EKUITAS', total_liab + total_ekuitas, r, color='BFDBFE'); r += 1
        if abs(selisih) > 1:
            ws3.cell(row=r, column=1, value=f'Selisih: {selisih:,.0f}').font = Font(color='DC2626', bold=True)

    except Exception as e:
        ws3.cell(row=1, column=1, value=f'Error membangun sheet Neraca: {e}').font = Font(color='DC2626', bold=True)

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 4: DAFTAR TRANSAKSI  (Nominal | Piutang | Hutang)
    # ══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Daftar Transaksi')
    ws4.column_dimensions['A'].width = 13
    ws4.column_dimensions['B'].width = 42
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 18
    ws4.column_dimensions['E'].width = 18
    ws4.column_dimensions['F'].width = 18

    _hdr1(ws4, 'DAFTAR TRANSAKSI', 1)
    ws4.cell(row=2, column=1,
             value=f'Periode: {sd}  s/d  {ed}').font = Font(italic=True, size=9, color='64748B')

    hdr_row = 4
    for col4, txt4 in enumerate(
            ['Tanggal', 'Keterangan', 'Jenis', 'Nominal (Rp)', 'Piutang (Rp)', 'Hutang (Rp)'], 1):
        ch = ws4.cell(row=hdr_row, column=col4, value=txt4)
        ch.font = Font(bold=True, color='FFFFFF', size=10)
        ch.fill = PatternFill(patternType='solid', fgColor='334155')
        ch.alignment = Alignment(horizontal='center')

    try:
        conn2 = db()

        # Pre-fetch piutang delta per jurnal (account 1120)
        piutang_map = {}
        for row_ in conn2.execute("""
            SELECT dj.jurnal_id, SUM(dj.debit) - SUM(dj.kredit) as delta
            FROM detail_jurnal dj
            JOIN akun a ON a.id = dj.akun_id
            WHERE a.kode = '1120'
            GROUP BY dj.jurnal_id
        """).fetchall():
            piutang_map[row_['jurnal_id']] = row_['delta'] or 0

        # Pre-fetch hutang delta per jurnal (account 2100)
        hutang_map = {}
        for row_ in conn2.execute("""
            SELECT dj.jurnal_id, SUM(dj.kredit) - SUM(dj.debit) as delta
            FROM detail_jurnal dj
            JOIN akun a ON a.id = dj.akun_id
            WHERE a.kode = '2100'
            GROUP BY dj.jurnal_id
        """).fetchall():
            hutang_map[row_['jurnal_id']] = row_['delta'] or 0

        # Main transaction rows
        tx_rows = conn2.execute("""
            SELECT j.id, j.tanggal, j.keterangan, j.tipe_tx, j.kategori,
                   COALESCE((
                       SELECT SUM(dj2.kredit)
                       FROM detail_jurnal dj2
                       JOIN akun a2 ON a2.id = dj2.akun_id
                       WHERE dj2.jurnal_id = j.id AND a2.kode LIKE '4%'
                   ), 0) as rev_kredit,
                   COALESCE((
                       SELECT SUM(d2.debit)
                       FROM detail_jurnal d2
                       WHERE d2.jurnal_id = j.id
                   ), 0) as total_debit
            FROM jurnal j
            WHERE j.tanggal >= ? AND j.tanggal <= ?
            ORDER BY j.tanggal, j.id
        """, (sd, ed)).fetchall()
        conn2.close()

        r4 = hdr_row + 1
        total_nominal4 = total_piutang4 = total_hutang4 = 0

        for tx in tx_rows:
            jid  = tx['id']
            pd4  = piutang_map.get(jid, 0)
            hd4  = hutang_map.get(jid, 0)
            is_masuk = tx['tipe_tx'] == 'PEMASUKAN'

            # Jenis label
            if pd4 > 0.005:
                jenis4 = 'Penjualan Kredit'
            elif pd4 < -0.005:
                jenis4 = 'Penerimaan Piutang'
            elif hd4 > 0.005:
                jenis4 = 'Pembelian Kredit'
            elif hd4 < -0.005:
                jenis4 = 'Pembayaran Hutang'
            elif is_masuk:
                jenis4 = 'Penjualan'
            elif tx['kategori'] == 'INVESTASI':
                jenis4 = 'Investasi Aset'
            elif tx['kategori'] == 'PENDANAAN':
                jenis4 = 'Penarikan Owner'
            else:
                jenis4 = 'Operasional'

            # Column values (mutually exclusive)
            if abs(pd4) > 0.005:
                nominal_v, piutang_v, hutang_v = None, pd4, None
                total_piutang4 += pd4
            elif abs(hd4) > 0.005:
                nominal_v, piutang_v, hutang_v = None, None, hd4
                total_hutang4 += hd4
            else:
                raw4 = (tx['rev_kredit'] if is_masuk else tx['total_debit']) or 0
                nominal_v = raw4 if is_masuk else -raw4
                piutang_v, hutang_v = None, None
                total_nominal4 += nominal_v

            ws4.cell(row=r4, column=1, value=str(tx['tanggal']))
            ws4.cell(row=r4, column=2, value=tx['keterangan'])
            ws4.cell(row=r4, column=3, value=jenis4)

            for col4v, val4 in [(4, nominal_v), (5, piutang_v), (6, hutang_v)]:
                if val4 is not None:
                    cv = ws4.cell(row=r4, column=col4v, value=val4)
                    cv.number_format = '#,##0'
                    cv.font = Font(color='16A34A' if val4 >= 0 else 'DC2626', bold=True)
            r4 += 1

        # Footer total row
        r4 += 1
        fill_foot = PatternFill(patternType='solid', fgColor='F1F5F9')
        ct = ws4.cell(row=r4, column=3, value='TOTAL')
        ct.font = Font(bold=True, size=10)
        ct.fill = fill_foot
        for col4v, val4 in [(4, total_nominal4), (5, total_piutang4), (6, total_hutang4)]:
            cv = ws4.cell(row=r4, column=col4v, value=val4)
            cv.number_format = '#,##0'
            cv.font = Font(bold=True, color='16A34A' if val4 >= 0 else 'DC2626')
            cv.fill = fill_foot

    except Exception as e4:
        ws4.cell(row=5, column=1,
                 value=f'Error Sheet 4: {e4}').font = Font(color='DC2626', bold=True)

    # ── Output ──────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"laporan_keuangan_{sd}_{ed}.xlsx"
    return send_file(buf, download_name=fname, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------- CLEAR DATA ----------
@app.route('/admin/clear-data', methods=['POST'])
@admin_required
def clear_data():
    konfirmasi = request.form.get('konfirmasi', '').strip()
    password   = request.form.get('password', '').strip()

    if konfirmasi != 'HAPUS SEMUA DATA':
        flash('Teks konfirmasi tidak sesuai. Data tidak dihapus.', 'danger')
        return redirect(url_for('settings'))

    conn = db()
    user = conn.execute("SELECT password_hash FROM users WHERE id=?", (session['user_id'],)).fetchone()
    conn.close()
    if not user or sha256(password.encode()).hexdigest() != user['password_hash']:
        flash('Password salah. Data tidak dihapus.', 'danger')
        return redirect(url_for('settings'))

    conn = db()
    conn.executescript("""
        DELETE FROM detail_jurnal;
        DELETE FROM jurnal;
        DELETE FROM bayar_piutang;
        DELETE FROM bayar_hutang;
        DELETE FROM piutang;
        DELETE FROM hutang;
        DELETE FROM invoice_item;
        DELETE FROM invoice;
        DELETE FROM pergerakan_stok;
        DELETE FROM hpp_produk;
        DELETE FROM produk;
        DELETE FROM aset_tetap;
        DELETE FROM log_aktivitas;
        DELETE FROM sqlite_sequence WHERE name IN
            ('jurnal','detail_jurnal','piutang','hutang','bayar_piutang','bayar_hutang',
             'invoice','invoice_item','produk','aset_tetap','pergerakan_stok','hpp_produk','log_aktivitas');
    """)
    reset_keys = ['nama_usaha','inv_nama','inv_tagline','inv_alamat','inv_email',
                  'inv_logo','inv_telepon','inv_catatan','inv_terms','inv_top_note',
                  'inv_rek','modal_awal']
    for k in reset_keys:
        conn.execute("UPDATE settings SET value='' WHERE key=?", (k,))
    conn.execute("UPDATE settings SET value='[]' WHERE key='inv_rek'")
    conn.commit(); conn.close()
    flash('Semua data berhasil dihapus. Aplikasi siap digunakan dari awal.', 'success')
    return redirect(url_for('dashboard'))


if __name__ == '__main__':
    init_db()
    print('\n' + '='*55)
    print('   FINANSIAL APP v4.0 - Sistem Akuntansi Lokal')
    print('   Buka browser: http://localhost:5000')
    print('   Login: admin / admin123')
    print('='*55 + '\n')
    app.run(debug=False, host='127.0.0.1', port=5000)
